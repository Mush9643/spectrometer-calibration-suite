import sys
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
import numpy as np
from PyQt6.QtCharts import QChart, QLineSeries, QValueAxis, QChartView, QLogValueAxis, QAbstractSeries, QScatterSeries
from PyQt6.QtWidgets import QApplication, QMainWindow, QPushButton, QVBoxLayout, QWidget, QMessageBox, QDialog, \
    QSplashScreen, QCheckBox, QLineEdit, QTabWidget, QListWidgetItem, QListWidget, QMenu, QHBoxLayout, QFileDialog, \
    QTableWidgetItem, QTableWidget, QHeaderView, QLabel
from PyQt6.QtGui import QPixmap, QIcon, QPainter, QDesktopServices, QColor, QFont
from PyQt6.QtCore import Qt, QTimer, QUrl, QRectF, pyqtSignal
from modbus import ModbusClient  # Импортируем ModbusClient из файла modbus.py
from settings_dialog import SettingsDialog  # Импортируем диалоговое окно настроек
from spectrum_addition import SpectrumAddition
import pandas as pd  # Для работы с Excel
from math_utils import highlight_am241_peak, add_recalculate_button
from math_utils import highlight_rn_peaks
from math_utils import add_calibration_button
from Beta_math import add_beta_calibration_button
from Beta_math import update_calibration_button_state
from fon_math import process_fon_data, process_isotope_data
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fon_math import NUD_b, VUD_b
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
import os
import logging
from gamma_math import print_gamma_impulses, calculate_peaks, plot_peaks, perform_calibration
from side_window_filler import SideWindow
from tkinter import filedialog, simpledialog, Tk
from datetime import datetime

##########################################################################
# Класс Зума
##########################################################################

class CustomChartView(QChartView):
    def __init__(self, chart, parent=None):
        super().__init__(chart, parent)
        self.setRenderHint(QPainter.RenderHint.Antialiasing)
        self._is_panning = False  # Флаг для режима перемещения
        self._last_pos = None  # Последняя позиция мыши для расчета смещения

    # В классе CustomChartView в main.py
    def wheelEvent(self, event):
        """Обработка прокрутки колесика мыши для зума относительно центра графика."""
        factor = 1.05 if event.angleDelta().y() > 0 else 0.95

        x_axis = self.chart().axes(Qt.Orientation.Horizontal)[0]
        y_axis = self.chart().axes(Qt.Orientation.Vertical)[0]
        x_min, x_max = x_axis.min(), x_axis.max()
        y_min, y_max = y_axis.min(), y_axis.max()

        new_x_range = (x_max - x_min) * factor
        new_y_range = (y_max - y_min) * factor

        if 10 < new_x_range < 2048:
            x_center = (x_min + x_max) / 2
            y_center = (y_min + y_max) / 2

            x_half_range = new_x_range / 2
            y_half_range = new_y_range / 2

            new_x_min = x_center - x_half_range
            new_x_max = x_center + x_half_range
            new_y_min = y_center - y_half_range
            new_y_max = y_center + y_half_range

            x_axis.setRange(new_x_min, new_x_max)
            y_axis.setRange(new_y_min, new_y_max)

            # Перерисовываем пики для Gamma
            if hasattr(self.parent(), 'reapply_gamma_peaks'):
                self.parent().reapply_gamma_peaks()

        event.accept()

    def mouseReleaseEvent(self, event):
        """Окончание перемещения."""
        if event.button() == Qt.MouseButton.LeftButton:
            self._is_panning = False
            self.setCursor(Qt.CursorShape.ArrowCursor)
            # Перерисовываем пики для Gamma после зума или перемещения
            if hasattr(self.parent(), 'reapply_gamma_peaks'):
                self.parent().reapply_gamma_peaks()
        super().mouseReleaseEvent(event)

    def mouseMoveEvent(self, event):
        """Перемещение графика при зажатой левой кнопке."""
        if self._is_panning and self._last_pos is not None:
            current_pos = event.position().toPoint()
            delta = current_pos - self._last_pos

            # Преобразование смещения пикселей в координаты графика
            x_axis = self.chart().axes(Qt.Orientation.Horizontal)[0]
            y_axis = self.chart().axes(Qt.Orientation.Vertical)[0]
            x_range = x_axis.max() - x_axis.min()
            y_range = y_axis.max() - y_axis.min()

            # Масштабирование смещения относительно размеров viewport
            dx = -delta.x() * x_range / self.viewport().width()
            dy = delta.y() * y_range / self.viewport().height()  # Инверсия для Y (вверх = уменьшение)

            # Обновление диапазонов осей
            x_axis.setRange(x_axis.min() + dx, x_axis.max() + dx)
            y_axis.setRange(y_axis.min() + dy, y_axis.max() + dy)

            self._last_pos = current_pos
            event.accept()
        else:
            super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
        """Окончание перемещения."""
        if event.button() == Qt.MouseButton.LeftButton:
            self._is_panning = False
            self.setCursor(Qt.CursorShape.ArrowCursor)  # Возвращаем стандартный курсор
        super().mouseReleaseEvent(event)

##########################################################################
# Класс Переключателя
##########################################################################

class ToggleSwitch(QWidget):
    stateChanged = pyqtSignal(int)  # Сигнал для изменения состояния

    def __init__(self, text="", parent=None):
        super().__init__(parent)
        self._checked = False
        self._text = text
        self.setFixedSize(265, 30)

    def text(self):  # Добавляем метод text()
        return self._text

    def setChecked(self, checked):
        if self._checked != checked:
            self._checked = checked
            self.update()  # Перерисовываем виджет
            self.stateChanged.emit(Qt.CheckState.Checked.value if checked else Qt.CheckState.Unchecked.value)

    def mousePressEvent(self, event):
        self.setChecked(not self._checked)

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)

        # Настройки шрифта для текста
        painter.setFont(QFont("Montserrat", 12))
        painter.setPen(QColor("#4A4A4A"))

        # Рисуем текст слева
        text_rect = QRectF(0, 0, 210, 30)  # Оставляем ширину текста без изменений
        painter.drawText(text_rect, Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter, self._text)

        # Рисуем фон переключателя справа, сдвигаем на 15px правее
        switch_rect = QRectF(215, 5, 40, 20)  # Сдвигаем с x=200 на x=215
        painter.setBrush(QColor("#4A4A4A"))  # Тёмно-серый фон
        painter.setPen(QColor("#000000"))  # Чёрный контур
        painter.drawRoundedRect(switch_rect, 10, 10)

        # Рисуем шар, корректируем позиции
        if self._checked:
            circle_pos = 237  # Позиция шара вправо (222 + 15 = 237)
            painter.setBrush(QColor("#C8102E"))  # Красный шар при активации
        else:
            circle_pos = 217  # Позиция шара влево (202 + 15 = 217)
            painter.setBrush(QColor("#FFFFFF"))  # Белый шар
        painter.setPen(Qt.PenStyle.NoPen)
        painter.drawEllipse(circle_pos, 7, 16, 16)

    def toggle(self):
        self.setChecked(not self._checked)

##########################################################################
# Класс Мелкого окна на вкладке menu
##########################################################################
class SpectrumGraphWindow(QMainWindow):
    def __init__(self, parent=None, modbus_client=None):
        super().__init__(parent)
        self.setWindowTitle("Спектр")
        self.setWindowIcon(QIcon(os.path.join("lib", "Pictures", "M-Photoroom.png")))
        self.resize(700, 500)
        self.modbus_client = modbus_client
        self.spectrum_values = []

        # Основной виджет и макет
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)

        # Создание графика
        self.chart = QChart()
        self.chart.setTitle("Спектр")
        title_font = QFont("Montserrat", 14, QFont.Weight.Bold)
        self.chart.setTitleFont(title_font)
        self.chart.setTitleBrush(QColor("#000000"))

        self.series = QLineSeries()
        self.series.setName("Импульсы")
        self.chart.addSeries(self.series)

        self.axis_x = QValueAxis()
        self.axis_x.setTitleText("Каналы")
        self.axis_x.setRange(0, 1023)
        axis_font = QFont("Montserrat", 12)
        self.axis_x.setLabelsFont(axis_font)
        self.axis_x.setTitleFont(axis_font)
        self.axis_x.setTitleBrush(QColor("#000000"))
        self.axis_x.setLabelsColor(QColor("#000000"))
        self.axis_x.setGridLineColor(QColor("#F5F5F5"))

        self.axis_y = QValueAxis()
        self.axis_y.setTitleText("Импульсы")
        self.axis_y.setRange(0, 200)  # Начальный диапазон
        self.axis_y.setLabelsFont(axis_font)
        self.axis_y.setTitleFont(axis_font)
        self.axis_y.setTitleBrush(QColor("#000000"))
        self.axis_y.setLabelsColor(QColor("#000000"))
        self.axis_y.setGridLineColor(QColor("#F5F5F5"))

        self.chart.addAxis(self.axis_x, Qt.AlignmentFlag.AlignBottom)
        self.chart.addAxis(self.axis_y, Qt.AlignmentFlag.AlignLeft)
        self.series.attachAxis(self.axis_x)
        self.series.attachAxis(self.axis_y)

        self.chart_view = CustomChartView(self.chart)
        self.chart_view.setRenderHint(QPainter.RenderHint.Antialiasing)
        layout.addWidget(self.chart_view)

        # Кнопка "Экспорт в Excel"
        self.export_button = QPushButton("Экспорт в Excel")
        self.export_button.setObjectName("exportButton")
        self.export_button.clicked.connect(self.export_to_excel)
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        button_layout.addWidget(self.export_button)
        button_layout.addStretch()
        layout.addLayout(button_layout)

        # Загружаем данные при открытии
        self.update_spectrum()

    def show_info_message(self, message):
        """Показывает информационное сообщение."""
        msg = QMessageBox(self)
        msg.setIcon(QMessageBox.Icon.Information)
        msg.setText(message)
        msg.setWindowTitle("Информация")
        msg.exec()

    def show_warning_message(self, message):
        """Показывает предупреждение."""
        msg = QMessageBox(self)
        msg.setIcon(QMessageBox.Icon.Warning)
        msg.setText(message)
        msg.setWindowTitle("Предупреждение")
        msg.exec()

    def show_error_message(self, message):
        """Показывает сообщение об ошибке."""
        msg = QMessageBox(self)
        msg.setIcon(QMessageBox.Icon.Critical)
        msg.setText(message)
        msg.setWindowTitle("Ошибка")
        msg.exec()

    def update_spectrum(self):
        """Обновляет спектр импульсов из Modbus."""
        if self.modbus_client:
            try:
                spectrum_values = self.modbus_client.read_spectrum(
                    start_register=0x0100,
                    num_registers=1024,
                    slave_address=1
                )
                self.spectrum_values = spectrum_values
                self.series.clear()
                for i, value in enumerate(spectrum_values):
                    display_value = 0 if i < 2 else value  # Первые два значения обнуляем
                    self.series.append(i, display_value)
                max_y = max(spectrum_values[2:], default=100)  # Пропускаем первые два значения
                self.axis_y.setRange(0, max_y * 1.1)
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Ошибка при чтении данных через Modbus: {str(e)}")
        else:
            QMessageBox.warning(self, "Предупреждение", "Modbus-клиент не инициализирован.")

    def export_to_excel(self):
        """Экспортирует данные спектра в Excel."""
        from PyQt6.QtWidgets import QInputDialog  # Добавляем импорт для QInputDialog
        try:
            if not hasattr(self, 'spectrum_values') or not self.spectrum_values:
                self.show_warning_message("Нет данных для экспорта.")
                return

            data = {
                "Точка спектра": list(range(len(self.spectrum_values))),
                "Значение": self.spectrum_values
            }
            df = pd.DataFrame(data)

            # Получаем путь к корневой директории и папке lib
            root_dir = os.path.dirname(os.path.abspath(__file__))
            lib_dir = os.path.join(root_dir, 'lib')
            os.makedirs(lib_dir, exist_ok=True)

            # Запрашиваем имя файла через QInputDialog
            custom_name, ok = QInputDialog.getText(self, "Имя файла", "Введите имя файла:")
            if not ok or not custom_name:
                return

            # Форматируем дату с нижними подчеркиваниями
            current_date = datetime.now().strftime('%Y_%m_%d')  # Например, 2025_04_07
            file_name = f"{custom_name}_{current_date}.xlsx"
            file_path = os.path.join(lib_dir, file_name)

            # Сохраняем файл
            df.to_excel(file_path, index=False, engine='openpyxl')
            self.show_info_message(f"Данные успешно экспортированы в {file_path}")

        except Exception as e:
            self.show_error_message(f"Ошибка при экспорте данных: {str(e)}")

##########################################################################
# Класс основного окна приложения
##########################################################################

class SpectrumWindow(QMainWindow):

    def __init__(self, modbus):
        """
        Инициализация основного окна приложения SpectrumWindow.
        """
        # =========================================================================
        # Блок 1: Инициализация базовых свойств и стилей
        # =========================================================================
        super().__init__()

        # Применение стилей с постельными тонами

        self.setStyleSheet("""
            /* Основные стили окна */
            QMainWindow {
                background-color: #FFFFFF; /* Белый фон */
                font-family: 'Montserrat', sans-serif; /* Устанавливаем шрифт Montserrat */
            }
            QTabWidget::pane {
                border: 1px solid #4A4A4A; /* Тёмно-серая граница */
                background-color: #FFFFFF; /* Белый фон */
                border-radius: 5px;
            }
            QTabBar::tab {
                background-color: #F5F5F5; /* Очень светлый серый для вкладок */
                color: #4A4A4A; /* Тёмно-серый текст */
                padding: 8px 16px;
                border-top-left-radius: 5px;
                border-top-right-radius: 5px;
                font-size: 14px;
                font-weight: 600; /* Montserrat SemiBold */
            }
            QTabBar::tab:selected {
                background-color: #FFFFFF; /* Белый фон для активной вкладки */
                color: #C0392B; /* Красный текст для активной вкладки (соответствует логотипу) */
                border-bottom: 2px solid #C0392B; /* Красный акцент */
            }
            QLineEdit {
                border: 1px solid #4A4A4A; /* Тёмно-серая граница */
                border-radius: 5px;
                padding: 5px;
                background-color: #FFFFFF; /* Белый фон */
                color: #000000; /* Чёрный текст */
                font-size: 12px;
                font-weight: 400; /* Montserrat Regular */
            }
            QLineEdit:read-only {
                background-color: #F5F5F5; /* Очень светлый серый для read-only */
            }
            QListWidget {
                background-color: #FFFFFF; /* Белый фон */
                border: 1px solid #4A4A4A; /* Тёмно-серая граница */
                border-radius: 5px;
                padding: 5px;
                color: #000000; /* Чёрный текст */
                font-size: 12px;
                font-weight: 400; /* Montserrat Regular */
            }
            QListWidget::item {
                padding: 5px;
            }
            QListWidget::item:selected {
                background-color: #C0392B; /* Красный фон для выбранного элемента (соответствует логотипу) */
                color: #FFFFFF; /* Белый текст */
            }
            QPushButton {
                background-color: #C0392B; /* Красный для основных кнопок (соответствует логотипу) */
                color: #FFFFFF; /* Белый текст */
                border: none;
                border-radius: 5px;
                padding: 8px 16px;
                font-size: 12px;
                font-weight: 600; /* Montserrat SemiBold */
            }
            QPushButton:hover {
                background-color: #A93226; /* Более тёмный красный при наведении */
            }
            QPushButton:pressed {
                background-color: #922B21; /* Ещё более тёмный красный при нажатии */
            }
            QPushButton#folderButton {
                background-color: transparent; /* Прозрачный фон для кнопки папки */
                border: none;
                padding: 2px;
                color: #4A4A4A; /* Тёмно-серый цвет текста */
                font-size: 16px; /* Размер символа */
            }
            QPushButton#folderButton:hover {
                background-color: rgba(200, 16, 46, 0.1); /* Лёгкий красный оттенок при наведении */
            }
            QPushButton#exportButton {
                background-color: #C0392B; /* Красный для кнопки экспорта */
                color: #FFFFFF; /* Белый текст */
            }
            QPushButton#exportButton:hover {
                background-color: #A93226;
            }
            QPushButton#exportButton:pressed {
                background-color: #922B21;
            }
            QPushButton#toggleCheckboxesButton {
                background-color: #4A4A4A; /* Тёмно-серый для вторичной кнопки */
                color: #FFFFFF; /* Белый текст */
            }
            QPushButton#toggleCheckboxesButton:hover {
                background-color: #5A5A5A; /* Светлее при наведении */
            }
            QPushButton#toggleCheckboxesButton:pressed {
                background-color: #3A3A3A; /* Темнее при нажатии */
            }
            QPushButton#resetZoomButton {
                background-color: #4A4A4A; /* Тёмно-серый для вторичной кнопки */
                color: #FFFFFF; /* Белый текст */
            }
            QPushButton#resetZoomButton:hover {
                background-color: #5A5A5A; /* Светлее при наведении */
            }
            QPushButton#resetZoomButton:pressed {
                background-color: #3A3A3A; /* Темнее при нажатии */
            }
            QWidget#checkboxesWidget {
                background-color: #FFFFFF; /* Белый фон */
                border: 1px solid #4A4A4A; /* Тёмно-серая граница */
                border-radius: 5px;
                padding: 5px;
            }
            QChartView {
                background-color: #FFFFFF; /* Белый фон графика */
            }
            /* Стили для таблицы на вкладке Report */
            QTableWidget {
                background-color: #F8FAFC; /* Светлый фон таблицы */
                border: 1px solid #4A4A4A; /* Тёмно-серая граница */
                border-radius: 5px;
                font-family: 'Montserrat', sans-serif;
                font-size: 12px;
            }
            QHeaderView::section {
                background-color: #C0392B; /* Красный фон для заголовка таблицы (соответствует логотипу) */
                color: #FFFFFF; /* Белый текст */
                padding: 5px;
                border: 1px solid #4A4A4A; /* Тёмно-серая граница */
                font-weight: 600; /* Montserrat SemiBold */
            }
            QTableWidget::item {
                padding: 5px;
            }
            QTableWidget::item:alternate {
                background-color: #F5F5F5; /* Чередование светло-серого фона */
            }
            /* Стили для кнопок на вкладке Report */
            QPushButton#refreshButton {
                background-color: #C0392B; /* Красный фон (соответствует логотипу) */
                color: #FFFFFF; /* Белый текст */
                border-radius: 5px;
                padding: 5px;
                max-width: 150px;
            }
            QPushButton#refreshButton:hover {
                background-color: #A93226;
            }
            QPushButton#refreshButton:pressed {
                background-color: #922B21;
            }
            QPushButton#reportButton {
                background-color: #C0392B; /* Красный фон (соответствует логотипу) */
                color: #FFFFFF; /* Белый текст */
                border-radius: 5px;
                padding: 5px;
                max-width: 150px;
            }
            QPushButton#reportButton:hover {
                background-color: #A93226;
            }
            QPushButton#reportButton:pressed {
                background-color: #922B21;
            }
            /* Стиль для меток (QLabel) */
            QLabel {
                color: #4A4A4A; /* Тёмно-серый текст для согласованности */
                font-size: 12px;
                padding: 5px;
            }
            QPushButton#sideWindowButton {
                background-color: #4A4A4A; /* Тёмно-серый фон */
                color: #FFFFFF; /* Белый текст */
                border-radius: 5px;
                padding: 5px;
                max-width: 150px;
            }
            QPushButton#sideWindowButton:hover {
                background-color: #5A5A5A; /* Светлее при наведении */
            }
            QPushButton#sideWindowButton:pressed {
                background-color: #3A3A3A; /* Темнее при нажатии */
            }
                /* Заголовок вкладки Combined Report */
            QLabel#tabTitle {
                color: #C0392B; /* Красный цвет, соответствующий брендингу */
                font-family: 'Montserrat', sans-serif;
                font-size: 14px;
                font-weight: 600; /* Montserrat SemiBold */
                padding: 10px;
            }

            /* Улучшенные стили для QListWidget с узорами для доступности */
            QListWidget#reportsList::item {
                padding: 5px;
            }
            QListWidget#reportsList::item:selected {
                background-color: #C0392B; /* Красный фон для выбранного элемента */
                color: #FFFFFF; /* Белый текст */
            }
            /* Цветовая кодировка с узорами для доступности */
            QListWidget#reportsList::item[category="spectral"] {
                background-color: #D6BCFA; /* Фиолетовый для спектральных данных */
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #D6BCFA, stop:1 #E9D8FD); /* Градиент с узором */
            }
            QListWidget#reportsList::item[category="report"] {
                background-color: #BBF7D0; /* Зелёный для отчётов */
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #BBF7D0, stop:1 #DCFCE7); /* Градиент с узором */
            }

            /* Стиль для метки статуса */
            QLabel#reportsStatusLabel {
                color: #4A4A4A; /* Тёмно-серый текст */
                font-size: 12px;
                padding: 5px;
                background-color: #F5F5F5; /* Светло-серый фон */
                border-radius: 5px;
            }

            /* Стиль для кнопки "Сборочный отчёт" */
            QPushButton#assemblyReportButton {
                background-color: #C0392B; /* Красный фон */
                color: #FFFFFF; /* Белый текст */
                border: none;
                border-radius: 5px;
                padding: 8px 16px;
                font-size: 12px;
                font-weight: 600; /* Montserrat SemiBold */
            }
            QPushButton#assemblyReportButton:hover {
                background-color: #A93226; /* Более тёмный красный при наведении */
            }
            QPushButton#assemblyReportButton:pressed {
                background-color: #922B21; /* Ещё более тёмный красный при нажатии */
            }
            /* Стили для кнопок на вкладке Report */
            QPushButton#reportButton {
                background-color: #C0392B; /* Красный фон */
                color: #FFFFFF; /* Белый текст */
                border: none;
                border-radius: 5px;
                padding: 10px 30px;
                font-size: 12px;
                font-weight: 600; /* Montserrat SemiBold */
            }
            QPushButton#reportButton:hover {
                background-color: #A93226; /* Более тёмный красный при наведении */
            }
            QPushButton#reportButton:pressed {
                background-color: #922B21; /* Ещё более тёмный красный при нажатии */
            }

            QPushButton#sideWindowButton {
                background-color: #C0392B; /* Красный фон (изменяем с тёмно-серого на красный для единообразия) */
                color: #FFFFFF; /* Белый текст */
                border: none;
                border-radius: 5px;
                padding: 10px 30px;
                font-size: 12px;
                font-weight: 600; /* Montserrat SemiBold */
            }
            QPushButton#sideWindowButton:hover {
                background-color: #A93226; /* Более тёмный красный при наведении */
            }
            QPushButton#sideWindowButton:pressed {
                background-color: #922B21; /* Ещё более тёмный красный при нажатии */
            }
            QPushButton#sideWindowButton {
                background-color: #C0392B; /* Красный фон (изменяем с тёмно-серого на красный для единообразия) */
                color: #FFFFFF; /* Белый текст */
                border: none;
                border-radius: 5px;
                padding: 10px 30px;
                font-size: 12px;
                font-weight: 600; /* Montserrat SemiBold */
            }
            QPushButton#sideWindowButton:hover {
                background-color: #A93226; /* Более тёмный красный при наведении */
            }
            QPushButton#sideWindowButton:pressed {
                background-color: #922B21; /* Ещё более тёмный красный при нажатии */
            }
            QPushButton#spectrumButton {
                background-color: #C0392B; /* Красный фон */
                color: #FFFFFF; /* Белый текст */
                border: none;
                border-radius: 5px;
                padding: 8px 16px;
                font-size: 12px;
                font-weight: 600; /* Montserrat SemiBold */
            }
            QPushButton#spectrumButton:hover {
                background-color: #A93226; /* Более тёмный красный при наведении */
            }
            QPushButton#spectrumButton:pressed {
                background-color: #922B21; /* Ещё более тёмный красный при нажатии */
            }
            QPushButton#alfaResetZoomButton {
                background-color: #4A4A4A; /* Средний серый фон */
                color: #FFFFFF; /* Белый текст */
                border: none;
                border-radius: 5px;
                padding: 8px 16px;
                font-size: 12px;
                font-weight: 600; /* Montserrat SemiBold */
            }
            QPushButton#alfaResetZoomButton:hover {
                background-color: #5A5A5A; /* Светло-серый при наведении */
            }
            QPushButton#alfaResetZoomButton:pressed {
                background-color: #3A3A3A; /* Тёмно-серый при нажатии */
            }
        """)

        # Настройка окна
        self.setWindowTitle("Спектр импульсов")
        self.setWindowIcon(QIcon("lib\\Pictures\\M-Photoroom.png"))

        # Инициализация данных
        self.fon_processed = False
        self.am241_data = []  # Массив для данных Am241
        self.c14_data = []  # Массив для данных C14
        self.cs137_data = []  # Массив для данных Cs137
        self.sry90_data = []  # Массив для данных SrY90
        self.rad_data = []  # Массив для данных Rad
        self.fon_data = []  # Массив для данных фона

        # Проверка и обработка modbus
        self.modbus_client = modbus  # Просто присваиваем, как в старой версии
        # Инициализация флага для отслеживания выполнения калибровки
        self.calibration_performed = False
        # =========================================================================
        # Блок 2: Инициализация словарей для хранения данных и цветов
        # =========================================================================
        # Словарь для хранения использованных цветов для Alfa и Beta
        self.used_alfa_colors = {}  # Ключ — имя файла, значение — QColor
        self.used_beta_colors = {}  # Ключ — имя файла, значение — QColor

        # Список уникальных цветов (RGB)
        self.available_colors = [
            QColor(255, 206, 86),  # Жёлтый
            QColor(75, 192, 192),  # Бирюзовый
            QColor(153, 102, 255),  # Фиолетовый
            QColor(255, 159, 64),  # Оранжевый
            QColor(0, 128, 0),  # Зелёный
            QColor(128, 0, 128),  # Пурпурный
            QColor(255, 165, 0),  # Золотой
            QColor(0, 191, 255),  # Глубокий голубой
            QColor(173, 216, 230),  # Светло-голубой
            QColor(135, 206, 235),  # Небесно-голубой
            QColor(240, 128, 128),  # Светло-розовый
            QColor(144, 238, 144),  # Салатовый
        ]

        # Словарь для хранения массивов данных Alfa
        self.alfa_data_arrays = {}
        self.first_impulse_values = {}

        # Инициализация серий для точек P90 (вертикальные линии) и пиков
        self.p90_series = {}
        self.calibration_coefficients = None
        self.peak_points = {}

        # Словари для хранения CheckBox
        self.alfa_checkboxes = {}
        self.beta_checkboxes = {}

        # Словари для хранения серий графиков
        self.alfa_series_dict = {}
        self.beta_series_dict = {}

        # =========================================================================
        # Блок 3: Создание вкладки "Menu"
        # =========================================================================
        # Вкладка "Menu"
        self.tabs = QTabWidget()
        self.setCentralWidget(self.tabs)

        # Вкладка "Menu"
        self.tab0 = QWidget()
        self.tabs.addTab(self.tab0, "Menu")
        menu_layout = QVBoxLayout()

        # Поле для ввода имени папки с кнопкой выбора папки
        folder_layout = QHBoxLayout()
        self.folder_input = QLineEdit("")  # По умолчанию папка "098"
        self.folder_input.setReadOnly(True)  # Запрещаем ручное редактирование
        folder_layout.addWidget(self.folder_input)

        # Кнопка с иконкой папки для выбора папки
        self.folder_button = QPushButton("📁")
        self.folder_button.setObjectName("folderButton")  # Для специфического стиля
        self.folder_button.setToolTip("Выберите папку с файлами")
        self.folder_button.clicked.connect(self.select_folder)  # Подключаем метод выбора папки
        folder_layout.addWidget(self.folder_button)

        menu_layout.addLayout(folder_layout)

        # Список для отображения файлов .xls
        self.file_list = QListWidget()
        self.file_list.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.file_list.customContextMenuRequested.connect(self.show_context_menu)  # Обработка правого клика
        menu_layout.addWidget(self.file_list)

        # Кнопка "Авто загрузка"
        self.auto_load_button = QPushButton("Авто загрузка")
        self.auto_load_button.setObjectName("autoLoadButton")  # Добавляем объектное имя для вторичного стиля
        self.auto_load_button.clicked.connect(self.auto_load_files)  # Подключаем метод для авто загрузки
        menu_layout.addWidget(self.auto_load_button)

        # Кнопка "Спектр"
        self.spectrum_button = QPushButton("Спектр")
        self.spectrum_button.setObjectName("spectrumButton")  # Для стилизации
        self.spectrum_button.clicked.connect(self.open_spectrum_window)  # Подключаем метод открытия окна
        menu_layout.addWidget(self.spectrum_button)

        self.tab0.setLayout(menu_layout)

        # Загружаем файлы .xls из папки по умолчанию
        self.load_xls_files()

        # =========================================================================
        # Блок 4: Создание вкладки "Alfa chart"
        # =========================================================================
        self.tab1 = QWidget()
        self.tabs.addTab(self.tab1, "Alfa")


        # Создание графика для вкладки Alfa
        self.chart = QChart()
        self.chart.setTitle("Alfa")
        title_font = QFont("Montserrat", 14, QFont.Weight.Bold)
        self.chart.setTitleFont(title_font)
        self.chart.setTitleBrush(QColor("#000000"))

        self.axis_x = QValueAxis()
        self.axis_x.setTitleText("Точка спектра")
        self.axis_x.setRange(0, 1023)
        self.axis_x.setTickCount(11)
        axis_font = QFont("Montserrat", 12)
        self.axis_x.setLabelsFont(axis_font)
        self.axis_x.setTitleFont(axis_font)
        self.axis_x.setTitleBrush(QColor("#000000"))
        self.axis_x.setLabelsColor(QColor("#000000"))
        self.axis_x.setGridLineColor(QColor("#F5F5F5"))

        self.axis_y = QValueAxis()
        self.axis_y.setTitleText("Значение импульса")
        self.axis_y.setRange(0, 1)  # Начальный пустой диапазон
        self.axis_y.setLabelFormat("%.2f")
        self.axis_y.setLabelsFont(axis_font)
        self.axis_y.setTitleFont(axis_font)
        self.axis_y.setTitleBrush(QColor("#000000"))
        self.axis_y.setLabelsColor(QColor("#000000"))
        self.axis_y.setGridLineColor(QColor("#F5F5F5"))

        self.chart.addAxis(self.axis_x, Qt.AlignmentFlag.AlignBottom)
        self.chart.addAxis(self.axis_y, Qt.AlignmentFlag.AlignLeft)

        self.chart_view = CustomChartView(self.chart)
        self.chart_view.setRenderHint(QPainter.RenderHint.Antialiasing)

        # Чекбокс для логарифмического масштаба
        self.log_checkbox = ToggleSwitch("Логарифмический масштаб")
        self.log_checkbox.stateChanged.connect(self.toggle_log_scale)


        # Создаем виджет для чекбоксов Alfa chart
        self.alfa_checkboxes_widget = QWidget()
        self.alfa_checkboxes_layout = QVBoxLayout()
        self.alfa_checkboxes_widget.setLayout(self.alfa_checkboxes_layout)
        self.alfa_checkboxes_widget.setObjectName("checkboxesWidget")
        self.alfa_checkboxes_widget.setMinimumHeight(50)
        self.alfa_checkboxes_widget.hide()

        # Кнопка для отображения/скрытия чекбоксов Alfa chart
        self.alfa_toggle_checkboxes_button = QPushButton("📋 Чекбоксы")
        self.alfa_toggle_checkboxes_button.setObjectName("toggleCheckboxesButton")
        self.alfa_toggle_checkboxes_button.clicked.connect(self.toggle_alfa_checkboxes)

        # Компоновка вкладки "Alfa chart"
        layout = QVBoxLayout()
        layout.addWidget(self.chart_view)

        controls_layout = QHBoxLayout()
        controls_layout.addWidget(self.log_checkbox)
        controls_layout.addStretch()
        controls_layout.addWidget(self.alfa_toggle_checkboxes_button)

        layout.addLayout(controls_layout)
        layout.addWidget(self.alfa_checkboxes_widget)

        self.tab1.setLayout(layout)

        add_calibration_button(self)
        self.add_reset_zoom_button()

        # =========================================================================
        # Блок 5: Создание вкладки "Beta chart"
        # =========================================================================
        # Вторая вкладка (Beta chart)
        self.tab2 = QWidget()
        self.tabs.addTab(self.tab2, "Beta")

        # Создание графика для второй вкладки
        self.beta_chart = QChart()
        self.beta_chart.setTitle("Beta")
        # Устанавливаем шрифт для заголовка
        title_font = QFont("Montserrat", 14, QFont.Weight.Bold)
        self.beta_chart.setTitleFont(title_font)
        self.beta_chart.setTitleBrush(QColor("#000000"))  # Чёрный цвет заголовка

        self.beta_series = QLineSeries()
        self.beta_series.setName("Beta данные")
        self.beta_chart.addSeries(self.beta_series)

        self.beta_axis_x = QValueAxis()
        self.beta_axis_x.setTitleText("Каналы")
        self.beta_axis_x.setRange(0, 100)
        # Устанавливаем шрифт и стиль для оси X
        axis_font = QFont("Montserrat", 12)
        self.beta_axis_x.setLabelsFont(axis_font)
        self.beta_axis_x.setTitleFont(axis_font)
        self.beta_axis_x.setTitleBrush(QColor("#000000"))  # Чёрный цвет текста
        self.beta_axis_x.setLabelsColor(QColor("#000000"))  # Чёрный цвет меток
        self.beta_axis_x.setGridLineColor(QColor("#F5F5F5"))  # Светло-серая сетка

        self.beta_axis_y = QValueAxis()
        self.beta_axis_y.setTitleText("Импульсы")
        self.beta_axis_y.setLabelsFont(axis_font)
        self.beta_axis_y.setTitleFont(axis_font)
        self.beta_axis_y.setTitleBrush(QColor("#000000"))  # Чёрный цвет текста
        self.beta_axis_y.setLabelsColor(QColor("#000000"))  # Чёрный цвет меток
        self.beta_axis_y.setGridLineColor(QColor("#F5F5F5"))  # Светло-серая сетка

        self.beta_chart.addAxis(self.beta_axis_x, Qt.AlignmentFlag.AlignBottom)
        self.beta_chart.addAxis(self.beta_axis_y, Qt.AlignmentFlag.AlignLeft)
        self.beta_series.attachAxis(self.beta_axis_x)
        self.beta_series.attachAxis(self.beta_axis_y)

        self.beta_chart_view = QChartView(self.beta_chart)
        self.beta_chart_view.setRenderHint(QPainter.RenderHint.Antialiasing)
        self.beta_chart_view.setRubberBand(QChartView.RubberBand.RectangleRubberBand)  # Включаем зум

        # Переключатель для логарифмического масштаба
        self.beta_log_checkbox = ToggleSwitch("Логарифмический масштаб")
        self.beta_log_checkbox.stateChanged.connect(self.toggle_beta_log_scale)

        # Создаем виджет для чекбоксов
        self.checkboxes_widget = QWidget()
        self.checkboxes_widget.setObjectName("checkboxesWidget")  # Для специфического стиля
        self.checkboxes_layout = QVBoxLayout()
        self.checkboxes_widget.setLayout(self.checkboxes_layout)
        self.checkboxes_widget.setMinimumHeight(50)  # Устанавливаем минимальную высоту
        self.checkboxes_widget.hide()  # Изначально скрыт

        # Кнопка для отображения/скрытия чекбоксов
        self.toggle_checkboxes_button = QPushButton("📋 Чекбоксы")
        self.toggle_checkboxes_button.setObjectName("toggleCheckboxesButton")
        self.toggle_checkboxes_button.clicked.connect(self.toggle_checkboxes)

        # Кнопка "Сбросить масштаб"
        self.beta_reset_zoom_button = QPushButton("Сбросить масштаб")
        self.beta_reset_zoom_button.setObjectName("resetZoomButton")
        self.beta_reset_zoom_button.clicked.connect(self.reset_beta_zoom)

        # Компоновка вкладки "Beta chart"
        beta_layout = QVBoxLayout()
        beta_layout.addWidget(self.beta_chart_view)

        # Горизонтальный layout для переключателя и кнопок
        controls_layout = QHBoxLayout()
        controls_layout.addWidget(self.beta_log_checkbox)
        controls_layout.addStretch()  # Растяжка для выравнивания
        controls_layout.addWidget(self.toggle_checkboxes_button)
        controls_layout.addWidget(self.beta_reset_zoom_button)
        beta_layout.addLayout(controls_layout)

        # Добавляем виджет с чекбоксами
        beta_layout.addWidget(self.checkboxes_widget)

        self.tab2.setLayout(beta_layout)

        # Добавляем кнопку калибровки
        add_beta_calibration_button(self)

        # =========================================================================
        # Блок 6: Создание вкладки "Gamma chart"
        # =========================================================================
        self.tab5 = QWidget()
        self.tabs.addTab(self.tab5, "Gamma")
        # Инициализация словаря для хранения пиков Gamma
        self.gamma_peaks = {}

        # Создание графика для вкладки Gamma
        self.gamma_chart = QChart()
        self.gamma_chart.setTitle("Gamma")
        # Устанавливаем шрифт для заголовка
        title_font = QFont("Montserrat", 14, QFont.Weight.Bold)
        self.gamma_chart.setTitleFont(title_font)
        self.gamma_chart.setTitleBrush(QColor("#000000"))  # Чёрный цвет заголовка

        self.gamma_series = QLineSeries()
        self.gamma_series.setName("Gamma данные")
        self.gamma_chart.addSeries(self.gamma_series)

        self.gamma_axis_x = QValueAxis()
        self.gamma_axis_x.setTitleText("Каналы")
        self.gamma_axis_x.setRange(500, 800)
        # Устанавливаем шрифт и стиль для оси X
        axis_font = QFont("Montserrat", 12)
        self.gamma_axis_x.setLabelsFont(axis_font)
        self.gamma_axis_x.setTitleFont(axis_font)
        self.gamma_axis_x.setTitleBrush(QColor("#000000"))  # Чёрный цвет текста
        self.gamma_axis_x.setLabelsColor(QColor("#000000"))  # Чёрный цвет меток
        self.gamma_axis_x.setGridLineColor(QColor("#F5F5F5"))  # Светло-серая сетка

        self.gamma_axis_y = QValueAxis()
        self.gamma_axis_y.setTitleText("Импульсы")
        self.gamma_axis_y.setLabelsFont(axis_font)
        self.gamma_axis_y.setTitleFont(axis_font)
        self.gamma_axis_y.setTitleBrush(QColor("#000000"))  # Чёрный цвет текста
        self.gamma_axis_y.setLabelsColor(QColor("#000000"))  # Чёрный цвет меток
        self.gamma_axis_y.setGridLineColor(QColor("#F5F5F5"))  # Светло-серая сетка

        self.gamma_chart.addAxis(self.gamma_axis_x, Qt.AlignmentFlag.AlignBottom)
        self.gamma_chart.addAxis(self.gamma_axis_y, Qt.AlignmentFlag.AlignLeft)
        self.gamma_series.attachAxis(self.gamma_axis_x)
        self.gamma_series.attachAxis(self.gamma_axis_y)

        self.gamma_chart_view = CustomChartView(self.gamma_chart)
        self.gamma_chart_view.setRenderHint(QPainter.RenderHint.Antialiasing)
        self.gamma_chart_view.setRubberBand(QChartView.RubberBand.RectangleRubberBand)  # Включаем зум

        # Переключатель для логарифмического масштаба
        self.gamma_log_checkbox = ToggleSwitch("Логарифмический масштаб")
        self.gamma_log_checkbox.stateChanged.connect(self.toggle_gamma_log_scale)

        # Создаем виджет для чекбоксов Gamma
        self.gamma_checkboxes_widget = QWidget()
        self.gamma_checkboxes_widget.setObjectName("checkboxesWidget")  # Для специфического стиля
        self.gamma_checkboxes_layout = QVBoxLayout()
        self.gamma_checkboxes_widget.setLayout(self.gamma_checkboxes_layout)
        self.gamma_checkboxes_widget.setMinimumHeight(50)  # Устанавливаем минимальную высоту
        self.gamma_checkboxes_widget.hide()  # Изначально скрыт

        # Кнопка для отображения/скрытия чекбоксов
        self.gamma_toggle_checkboxes_button = QPushButton("📋 Чекбоксы")
        self.gamma_toggle_checkboxes_button.setObjectName("toggleCheckboxesButton")
        self.gamma_toggle_checkboxes_button.clicked.connect(self.toggle_gamma_checkboxes)

        # Кнопка "Сбросить масштаб"
        self.gamma_reset_zoom_button = QPushButton("Сбросить масштаб")
        self.gamma_reset_zoom_button.setObjectName("resetZoomButton")
        self.gamma_reset_zoom_button.clicked.connect(self.reset_gamma_zoom)

        # Компоновка вкладки "Gamma chart"
        gamma_layout = QVBoxLayout()
        gamma_layout.addWidget(self.gamma_chart_view)

        # Горизонтальный layout для переключателя и кнопок
        gamma_controls_layout = QHBoxLayout()
        gamma_controls_layout.addWidget(self.gamma_log_checkbox)
        gamma_controls_layout.addStretch()  # Растяжка для выравнивания
        gamma_controls_layout.addWidget(self.gamma_toggle_checkboxes_button)
        gamma_controls_layout.addWidget(self.gamma_reset_zoom_button)
        gamma_layout.addLayout(gamma_controls_layout)

        # Добавляем виджет с чекбоксами
        gamma_layout.addWidget(self.gamma_checkboxes_widget)

        self.tab5.setLayout(gamma_layout)

        # Инициализация словарей для Gamma
        self.gamma_series_dict = {}  # Словарь для хранения серий Gamma
        self.gamma_checkboxes = {}  # Словарь для хранения чекбоксов Gamma
        self.used_gamma_colors = {}  # Словарь для хранения использованных цветов Gamma

        # =========================================================================
        # Блок 7: Создание вкладки "Report"
        # =========================================================================
        self.tab4 = QWidget()
        self.tabs.addTab(self.tab4, "Report")

        calibration_layout = QVBoxLayout()

        self.calibration_table = QTableWidget()
        self.calibration_table.setColumnCount(2)
        self.calibration_table.setRowCount(17)  # Устанавливаем количество строк
        self.calibration_table.setHorizontalHeaderLabels(["Параметр", "Значение"])


        data = [
            ("ИМЯ ПАКЛИ", "0000"),
            ("a (Alfa)", "Отсутствует"),
            ("b (Alfa)", "Отсутствует"),
            ("НУЛ α, № канала", "Отсутствует"),
            ("БУЛ ROI 3, № канала", "Отсутствует"),
            ("БУЛ ROI 2, № канала", "Отсутствует"),
            ("БУЛ ROI 6, № канала", "Отсутствует"),
            ("БУЛ ROI 4, № канала", "Отсутствует"),
            ("БУЛ ROI 5, № канала", "Отсутствует"),
            ("K(Po218)", "Отсутствует"),
            ("k1p9", "Отсутствует"),
            ("Пик Am241", "Отсутствует"),
            ("НУЛ β", "Отсутствует"),
            ("БУЛ β", "Отсутствует"),
            ("k1c0", "Отсутствует"),
            ("a (Beta)", "Отсутствует"),
            ("b (Beta)", "Отсутствует"),
        ]

        for row, (param, value) in enumerate(data):
            self.calibration_table.setItem(row, 0, QTableWidgetItem(param))
            item = QTableWidgetItem(value)
            if value == "Отсутствует":
                item.setForeground(QColor("#C0392B"))  # Красный цвет для "Отсутствует"
            self.calibration_table.setItem(row, 1, item)

        self.calibration_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.calibration_table.setAlternatingRowColors(True)  # Включаем чередование цветов строк
        calibration_layout.addWidget(self.calibration_table)

        # Горизонтальный layout для кнопок
        buttons_layout = QHBoxLayout()
        buttons_layout.addStretch()

        # Кнопка "Создание отчёта"
        self.report_button = QPushButton("Создание отчёта")
        self.report_button.setObjectName("reportButton")
        self.report_button.clicked.connect(self.generate_report)
        buttons_layout.addWidget(self.report_button)

        # Кнопка "Открыть доп. окно"
        self.side_window_button = QPushButton("Запись в детектор")
        self.side_window_button.setObjectName("sideWindowButton")
        self.side_window_button.clicked.connect(self.open_side_window)
        buttons_layout.addWidget(self.side_window_button)

        buttons_layout.addStretch()

        calibration_layout.addLayout(buttons_layout)

        self.tab4.setLayout(calibration_layout)

        add_recalculate_button(self)
        self.use_three_peaks = True

        # =========================================================================
        # Блок 8: Создание вкладки "Combined Report"
        # =========================================================================
        self.tab3 = QWidget()
        self.tabs.addTab(self.tab3, "Combined Report")

        # Добавляем заголовок вкладки
        self.combined_report_title = QLabel("Combined Report")
        self.combined_report_title.setObjectName("tabTitle")
        self.combined_report_title.setAlignment(Qt.AlignmentFlag.AlignCenter)

        # Список файлов отчётов
        self.reports_list = QListWidget()
        self.reports_list.setObjectName("reportsList")

        # Устанавливаем цветовую кодировку с категориями
        def update_list_item_styles():
            for i in range(self.reports_list.count()):
                item = self.reports_list.item(i)
                file_name = item.text().lower()
                if "отчёт" in file_name:
                    item.setData(Qt.ItemDataRole.UserRole, "report")  # Категория "report"
                else:
                    item.setData(Qt.ItemDataRole.UserRole, "spectral")  # Категория "spectral"

        self.reports_list.itemChanged.connect(update_list_item_styles)

        # Метка статуса
        self.reports_status_label = QLabel("Проверка наличия папки 'Отчёты'...")
        self.reports_status_label.setObjectName("reportsStatusLabel")
        self.reports_status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

        # Кнопка "Сборочный отчёт"
        self.assembly_report_button = QPushButton("Сборочный отчёт")
        self.assembly_report_button.setObjectName("assemblyReportButton")
        self.assembly_report_button.clicked.connect(self.create_assembly_report)
        self.assembly_report_button.setFixedWidth(200)
        self.assembly_report_button.setToolTip("Создать сборочный отчёт из выбранных файлов")

        # Макет для метки статуса
        status_layout = QHBoxLayout()
        status_layout.addStretch()
        status_layout.addWidget(self.reports_status_label)
        status_layout.addStretch()

        # Макет для кнопки
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        button_layout.addWidget(self.assembly_report_button)
        button_layout.addStretch()

        # Основной макет вкладки
        combined_layout = QVBoxLayout()
        combined_layout.addWidget(self.combined_report_title)  # Добавляем заголовок
        combined_layout.addLayout(status_layout)
        combined_layout.addWidget(self.reports_list)
        combined_layout.addLayout(button_layout)
        combined_layout.setSpacing(10)
        combined_layout.setContentsMargins(20, 20, 20, 20)

        self.tab3.setLayout(combined_layout)

        self.tabs.currentChanged.connect(self.on_tab_changed)
        self.reports_list.itemDoubleClicked.connect(self.open_report_file)

        # Обновляем стили списка после его заполнения
        self.reports_list.itemChanged.connect(update_list_item_styles)

        # =========================================================================
        # Блок 9: Добавление кнопки "Калибровка"
        # =========================================================================
        self.calibration_button = QPushButton("Калибровка")
        self.calibration_button.setObjectName("calibrationButton")  # Для стилизации через CSS
        self.calibration_button.clicked.connect(self.perform_calibration)  # Подключаем метод калибровки
        self.calibration_button.setFixedHeight(50)  # Устанавливаем высоту кнопки
        self.calibration_button.setFont(QFont("Montserrat", 14, QFont.Weight.Bold))  # Устанавливаем шрифт

        # Добавляем кнопку в основной layout
        gamma_layout.addWidget(self.calibration_button)

    ##########################################################################
    # Методы всякого разного
    ##########################################################################

    def open_spectrum_window(self):
        """Открывает окно с графиком спектра."""
        self.spectrum_window = SpectrumGraphWindow(self, self.modbus_client)
        self.spectrum_window.show()

    def open_report_file(self, item):
        """Открывает выбранный файл Excel."""
        reports_dir = Path("Отчёты")
        file_path = reports_dir / item.text()

        if not file_path.exists():
            self.reports_status_label.setText(f"Файл {item.text()} не найден")
            return

        try:
            if os.name == "nt":  # Windows
                os.startfile(file_path)
            else:  # macOS/Linux
                import subprocess
                subprocess.run(["xdg-open", str(file_path)])
            self.reports_status_label.setText(f"Открыт файл: {item.text()}")
        except Exception as e:
            self.reports_status_label.setText(f"Ошибка при открытии файла: {str(e)}")

    def create_assembly_reports_folder(self):
        """Создает папку 'Сборочные отчёты' в корневой директории приложения."""
        assembly_reports_dir = Path("Сборочные отчёты")
        try:
            if not assembly_reports_dir.exists():
                assembly_reports_dir.mkdir()
        except Exception as e:
            self.reports_status_label.setText(f"Ошибка при создании папки: {str(e)}")

    def create_assembly_report(self):
        """Создает сборочный отчёт с одним столбцом 'Параметр' и значениями по папкам."""
        # Показываем, что процесс начался
        self.reports_status_label.setText("Создание сборочного отчёта...")
        QApplication.processEvents()  # Обновляем интерфейс

        assembly_reports_dir = Path("Сборочные отчёты")
        try:
            if not assembly_reports_dir.exists():
                assembly_reports_dir.mkdir()
                self.reports_status_label.setText("Папка 'Сборочные отчёты' создана")
        except Exception as e:
            self.reports_status_label.setText(f"Ошибка при создании папки: {str(e)}")
            return

        displayed_files = [self.reports_list.item(i).text() for i in range(self.reports_list.count())]
        if not displayed_files:
            self.reports_status_label.setText("Нет файлов для создания сборочного отчёта")
            return

        # Список параметров в нужном порядке
        parameters_order = [
            "Имя папки",
            "a (Alfa)",
            "b (Alfa)",
            "НУД α, № канала (2700)",
            "ВУД ROI 3, № канала (4385.6)",
            "НУД ROI 2, № канала (5687.5)",
            "НУД ROI 6, № канала (6192.35)",
            "НУД ROI 4, № канала (6337.7)",
            "НУД ROI 5, № канала (8044.6)",
            "K(Po218)",
            "k1p9",
            "Пик Am241",
            "НУД β",
            "ВУД β",
            "k1c0",
            "a (Beta)",
            "b (Beta)",
            "Pn (80 кэВ)",
            "Pn (146 кэВ)",
            "Pn (400 кэВ)",
            "Pn (850 кэВ)",
            "Pn (1500 кэВ)",
            "Pn (2515 кэВ)"
        ]

        # Чтение данных и создание словаря с данными по папкам
        reports_dir = Path("Отчёты")
        data_by_folder = {}
        for file_name in displayed_files:
            file_path = reports_dir / file_name
            if not file_path.exists():
                self.reports_status_label.setText(f"Файл {file_name} не найден")
                continue

            try:
                df = pd.read_excel(file_path)
                if df.shape[1] != 2 or list(df.columns) != ["Параметр", "Значение"]:
                    self.reports_status_label.setText(f"Файл {file_name} имеет неверную структуру")
                    return
                folder_name = df[df["Параметр"] == "Имя папки"]["Значение"].iloc[0]
                data_dict = dict(zip(df["Параметр"], df["Значение"]))
                data_by_folder[folder_name] = data_dict
            except Exception as e:
                self.reports_status_label.setText(f"Ошибка при чтении файла {file_name}: {str(e)}")
                continue

        if not data_by_folder:
            self.reports_status_label.setText("Не удалось прочитать ни один файл")
            return

        # Создание итогового DataFrame
        combined_df = pd.DataFrame({"Параметр": parameters_order})
        for folder_name, data_dict in data_by_folder.items():
            values = [data_dict.get(param, "Отсутствует") for param in parameters_order]
            combined_df[folder_name] = values

        # Сохранение в Excel
        date_str = datetime.now().strftime("%Y-%m-%d")
        output_file = assembly_reports_dir / f"Сборочный отчёт_{date_str}.xlsx"
        combined_df.to_excel(output_file, index=False)

        # Форматирование с помощью openpyxl
        workbook = load_workbook(output_file)
        worksheet = workbook.active

        header_fill = PatternFill(start_color="D7CCC8", end_color="D7CCC8", fill_type="solid")
        data_fill = PatternFill(start_color="E6EAD6", end_color="E6EAD6", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        thick_border = Border(right=Side(style='medium', color='000000'))

        for cell in worksheet[1]:
            cell.fill = header_fill

        for row in range(2, worksheet.max_row + 1):
            for cell in worksheet[row]:
                cell.fill = data_fill

        for row in worksheet.rows:
            for cell in row:
                cell.border = thin_border

        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            worksheet.column_dimensions[column_letter].width = adjusted_width

        for row in range(1, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=1)
            new_border = Border(
                left=cell.border.left,
                right=thick_border.right,
                top=cell.border.top,
                bottom=cell.border.bottom
            )
            cell.border = new_border

        workbook.save(output_file)
        self.reports_status_label.setText(f"Сборочный отчёт создан: {output_file.name}")
        if os.name == "nt":
            os.startfile(output_file)
        else:
            import subprocess
            subprocess.run(["xdg-open", str(output_file)])

    def update_reports_list(self, index):
        """Обновляет список файлов Excel на вкладке Combined Report при переключении вкладок."""
        if index != self.tabs.indexOf(self.tab3):
            return

        self.reports_list.clear()

        reports_dir = Path("Отчёты")
        if not reports_dir.exists() or not reports_dir.is_dir():
            self.reports_status_label.setText("Папка 'Отчёты' не найдена")
            return

        excel_files = []
        for file_path in reports_dir.glob("*"):
            if file_path.is_file() and file_path.suffix.lower() in (".xlsx", ".xls"):
                excel_files.append(file_path)

        if not excel_files:
            self.reports_status_label.setText("Папка 'Отчёты' пуста или не содержит файлов Excel")
            return

        self.reports_status_label.setText(f"Найдено файлов Excel: {len(excel_files)}")
        for file_path in excel_files:
            self.reports_list.addItem(file_path.name)

    def reset_beta_zoom(self):
        """Сбрасывает масштаб графика Beta до исходного состояния."""
        self.beta_axis_x.setRange(0, 100)
        if self.beta_series_dict:
            max_y = max(
                max(series.points(), key=lambda p: p.y()).y()
                for series in self.beta_series_dict.values()
                if series.points()
            )
            self.beta_axis_y.setRange(0, max_y * 1.1)
        else:
            self.beta_axis_y.setRange(0, 1)
        self.beta_chart_view.update()

    def add_beta_reset_zoom_button(self):
        """Добавляет кнопку сброса масштаба на вкладку Beta chart в beta_controls_layout."""
        reset_button = QPushButton("Сбросить масштаб")
        reset_button.clicked.connect(self.reset_beta_zoom)
        main_layout = self.tab2.layout()
        controls_layout = None
        for i in range(main_layout.count()):
            item = main_layout.itemAt(i)
            if isinstance(item.layout(), QHBoxLayout):
                controls_layout = item.layout()
                break
        if controls_layout:
            controls_layout.addWidget(reset_button)
        else:
            new_layout = QHBoxLayout()
            new_layout.addWidget(self.beta_log_checkbox if hasattr(self, 'beta_log_checkbox') else QWidget())
            new_layout.addStretch()
            new_layout.addWidget(reset_button)
            main_layout.addLayout(new_layout)

    def reset_zoom(self):
        """Сбрасывает масштаб графика Alfa до исходного состояния."""
        # Устанавливаем исходный диапазон по X (обычно 0-1023 для каналов)
        self.axis_x.setRange(0, 1023)

        # Восстанавливаем диапазон по Y на основе данных
        if self.alfa_series_dict:
            max_y = max(
                max(series.points(), key=lambda p: p.y()).y()
                for series in self.alfa_series_dict.values()
                if series.points()
            )
            self.axis_y.setRange(0, max_y * 1.1)  # Добавляем 10% сверху для видимости
        else:
            self.axis_y.setRange(0, 1)  # Если данных нет, минимальный диапазон

        # Обновляем график
        self.chart_view.update()

    def add_reset_zoom_button(self):
        """Добавляет кнопку сброса масштаба на вкладку Alfa chart в controls_layout."""
        reset_button = QPushButton("Сбросить масштаб")
        reset_button.setObjectName("alfaResetZoomButton")
        reset_button.clicked.connect(self.reset_zoom)

        # Ищем controls_layout в компоновке tab1
        main_layout = self.tab1.layout()
        controls_layout = None
        for i in range(main_layout.count()):
            item = main_layout.itemAt(i)
            if isinstance(item.layout(), QHBoxLayout):
                controls_layout = item.layout()
                break

        if controls_layout:
            controls_layout.addWidget(reset_button)
        else:
            # Если controls_layout не найден, создаем новый
            new_layout = QHBoxLayout()
            new_layout.addWidget(self.log_checkbox if hasattr(self, 'log_checkbox') else QWidget())
            new_layout.addStretch()
            new_layout.addWidget(reset_button)
            main_layout.addLayout(new_layout)

    def update_calibration_table(self):
        """Обновляет таблицу с параметрами калибровки"""
        # Список параметров и их значений (аналогично generate_report)
        parameters = []

        # folder_name
        parameters.append(("Имя папки", self.folder_input.text()))

        # aRa и bRa (Alfa)
        if hasattr(self, 'calibration_coefficients') and self.calibration_coefficients:
            aRa, bRa = self.calibration_coefficients
            parameters.append(("a (Alfa)", f"{aRa:.3f}"))
            parameters.append(("b (Alfa)", f"{bRa:.3f}"))
        else:
            parameters.append(("a (Alfa)", "Отсутствует"))
            parameters.append(("b (Alfa)", "Отсутствует"))

        # Значения P()
        p_names = [
            "НУД α, № канала", "ВУД ROI 3, № канала", "НУД ROI 2, № канала",
            "НУД ROI 6, № канала", "НУД ROI 4, № канала", "НУД ROI 5, № канала"
        ]
        if hasattr(self, 'p_values') and self.p_values:
            p_values_list = list(self.p_values.items())
            for i in range(6):
                if i < len(p_values_list):
                    energy, p_value = p_values_list[i]
                    parameters.append((f"{p_names[i]} ({energy})", f"{p_value}"))
                else:
                    parameters.append((p_names[i], "Данные отсутствуют"))
        else:
            for name in p_names:
                parameters.append((name, "Отсутствует"))

        # K(Po218)
        if hasattr(self, 'ra_value') and self.ra_value is not None:
            parameters.append(("K(Po218)", f"{self.ra_value:.3f}"))
        else:
            parameters.append(("K(Po218)", "Отсутствует"))

        # k1p9
        if hasattr(self, 'k1p9_value') and self.k1p9_value is not None:
            parameters.append(("k1p9", f"{self.k1p9_value:.3f}"))
        else:
            parameters.append(("k1p9", "Отсутствует"))

        # Пик Am241
        if hasattr(self, 'peak_points') and self.peak_points:
            for series_name, peak in self.peak_points.items():
                if "Am241" in series_name and peak and peak.points():
                    am241_x = peak.points()[0].x()
                    parameters.append(("Пик Am241", f"{am241_x:.3f}"))
                    break
            else:
                parameters.append(("Пик Am241", "Не найден"))
        else:
            parameters.append(("Пик Am241", "Отсутствует"))

        # НУД β и ВУД β
        parameters.append(("НУД β", f"{NUD_b:.3f}"))
        parameters.append(("ВУД β", f"{VUD_b:.3f}"))

        # k1c0
        if hasattr(self, 'k1c0') and self.k1c0 is not None:
            parameters.append(("k1c0", f"{self.k1c0:.3f}"))
        else:
            parameters.append(("k1c0", "Отсутствует"))

        # AB и BB (Beta)
        if hasattr(self, 'beta_calibration_coefficients') and self.beta_calibration_coefficients:
            intercept, slope = self.beta_calibration_coefficients
            parameters.append(("a (Beta)", f"{intercept:.3f}"))
            parameters.append(("b (Beta)", f"{slope:.3f}"))
        else:
            parameters.append(("a (Beta)", "Отсутствует"))
            parameters.append(("b (Beta)", "Отсутствует"))

        # Добавляем значения Pn для Gamma
        Ep = [80, 146, 400, 850, 1500, 2515]  # Энергии из gamma_math.py
        if hasattr(self, 'gamma_pn_values') and self.gamma_pn_values:
            for i, (energy, pn_value) in enumerate(zip(Ep, self.gamma_pn_values)):
                parameters.append((f"Pn ({energy} кэВ)", f"{pn_value}"))
        else:
            for energy in Ep:
                parameters.append((f"Pn ({energy} кэВ)", "Отсутствует"))

        # Устанавливаем количество строк в таблице
        self.calibration_table.setRowCount(len(parameters))

        # Заполняем таблицу
        for row, (param, value) in enumerate(parameters):
            param_item = QTableWidgetItem(param)
            value_item = QTableWidgetItem(value)

            # Выравнивание по центру
            param_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            value_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

            # Запрещаем редактирование
            param_item.setFlags(param_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            value_item.setFlags(value_item.flags() & ~Qt.ItemFlag.ItemIsEditable)

            self.calibration_table.setItem(row, 0, param_item)
            self.calibration_table.setItem(row, 1, value_item)

    def select_folder(self):
        """Открывает диалог выбора папки и загружает файлы."""
        folder = QFileDialog.getExistingDirectory(self, "Выберите папку", os.getcwd())
        if folder:
            self.folder_input.setText(os.path.basename(folder))
            self.reset_all_data()
            self.load_xls_files()

    def clear_alfa_chart(self):
        """Очищает график Alfa chart."""
        self.series.clear()
        self.axis_y.setRange(0, 1)  # Сбрасываем диапазон оси Y
        self.chart_view.update()

    def toggle_checkboxes(self):
        """Переключает видимость виджета с чекбоксами и изменяет размер окна."""
        if self.checkboxes_widget.isVisible():
            # Скрываем виджет с чекбоксами
            self.checkboxes_widget.hide()
            self.toggle_checkboxes_button.setText("📋 Чекбоксы")
            # Возвращаем исходный размер окна
            self.resize(self.width(), self.original_height)
        else:
            # Показываем виджет с чекбоксами
            self.checkboxes_widget.show()
            self.toggle_checkboxes_button.setText("📋 Скрыть")
            # Увеличиваем высоту окна на размер виджета с чекбоксами
            checkboxes_height = self.checkboxes_widget.sizeHint().height()
            new_height = self.height() + checkboxes_height
            self.resize(self.width(), new_height)

    def toggle_alfa_checkboxes(self):
        """Переключает видимость виджета с чекбоксами для Alfa chart и изменяет размер окна."""
        if self.alfa_checkboxes_widget.isVisible():
            # Скрываем виджет с чекбоксами
            self.alfa_checkboxes_widget.hide()
            self.alfa_toggle_checkboxes_button.setText("📋 Чекбоксы")
            # Возвращаем исходный размер окна
            self.resize(self.width(), self.original_height)
        else:
            # Показываем виджет с чекбоксами
            self.alfa_checkboxes_widget.show()
            self.alfa_toggle_checkboxes_button.setText("📋 Скрыть")
            # Увеличиваем высоту окна на размер виджета с чекбоксами
            checkboxes_height = self.alfa_checkboxes_widget.sizeHint().height()
            new_height = self.height() + checkboxes_height
            self.resize(self.width(), new_height)

    def generate_report(self):
        """Метод для создания отчёта: создаёт Excel-файл в папке 'Отчёты' с параметрами, включая массив Pn."""
        # Проверяем, была ли выполнена калибровка (оставляем как есть для Alfa chart)
        if not self.calibration_performed:
            self.show_warning_message("Сначала выполните калибровку на вкладке Alfa chart.")
            print("Предупреждение: Калибровка не выполнена. Нажмите кнопку 'Калибровка' на вкладке Alfa chart.")
            return

        # Получаем имя папки из поля ввода
        folder_name = self.folder_input.text()

        # Создаём путь к папке "Отчёты" в корневой директории приложения
        reports_dir = os.path.join(os.getcwd(), "Отчёты")
        if not os.path.exists(reports_dir):
            os.makedirs(reports_dir)  # Создаём папку, если её нет

        # Формируем имя файла
        report_filename = f"{folder_name}_Отчёт.xlsx"
        report_path = os.path.join(reports_dir, report_filename)

        # Собираем данные для отчёта
        report_data = []

        # Заголовок
        report_data.append(["Параметр", "Значение"])

        # folder_name
        report_data.append(["Имя папки", folder_name])

        # aRa и bRa (Alfa)
        if hasattr(self, 'calibration_coefficients') and self.calibration_coefficients:
            aRa, bRa = self.calibration_coefficients
            report_data.append(["a (Alfa)", f"{aRa:.3f}"])
            report_data.append(["b (Alfa)", f"{bRa:.3f}"])
        else:
            report_data.append(["a (Alfa)", "Отсутствует"])
            report_data.append(["b (Alfa)", "Отсутствует"])

        # Значения P()
        if hasattr(self, 'p_values') and self.p_values:
            p_names = ["НУД α, № канала", "ВУД ROI 3, № канала", "НУД ROI 2, № канала",
                       "НУД ROI 6, № канала", "НУД ROI 4, № канала", "НУД ROI 5, № канала"]
            p_values_list = list(self.p_values.items())
            for i in range(6):
                if i < len(p_values_list):
                    energy, p_value = p_values_list[i]
                    report_data.append([f"{p_names[i]} ({energy})", f"{p_value}"])
                else:
                    report_data.append([p_names[i], "Данные отсутствуют"])
        else:
            for name in ["НУД α, № канала", "ВУД ROI 3, № канала", "НУД ROI 2, № канала",
                         "НУД ROI 6, № канала", "НУД ROI 4, № канала", "НУД ROI 5, № канала"]:
                report_data.append([name, "Отсутствует"])

        # K(Po218) (ранее ra)
        if hasattr(self, 'ra_value') and self.ra_value is not None:
            report_data.append(["K(Po218)", f"{self.ra_value:.3f}"])
        else:
            report_data.append(["K(Po218)", "Отсутствует"])

        # k1p9
        if hasattr(self, 'k1p9_value') and self.k1p9_value is not None:
            report_data.append(["k1p9", f"{self.k1p9_value:.3f}"])
        else:
            report_data.append(["k1p9", "Отсутствует"])

        # Пик Am241
        if hasattr(self, 'peak_points') and self.peak_points:
            for series_name, peak in self.peak_points.items():
                if "Am241" in series_name and peak and peak.points():
                    am241_x = peak.points()[0].x()
                    report_data.append(["Пик Am241", f"{am241_x:.3f}"])
                    break
            else:
                report_data.append(["Пик Am241", "Не найден"])
        else:
            report_data.append(["Пик Am241", "Отсутствует"])

        # НУД β и ВУД β
        report_data.append(["НУД β", f"{NUD_b:.3f}"])
        report_data.append(["ВУД β", f"{VUD_b:.3f}"])

        # k1c0
        if hasattr(self, 'k1c0') and self.k1c0 is not None:
            report_data.append(["k1c0", f"{self.k1c0:.3f}"])
        else:
            report_data.append(["k1c0", "Отсутствует"])

        # AB и BB (Beta) переименованы в intercept и slope
        if hasattr(self, 'beta_calibration_coequettes') and self.beta_calibration_coefficients:
            intercept, slope = self.beta_calibration_coefficients
            report_data.append(["a (Beta)", f"{intercept:.3f}"])  # intercept как a
            report_data.append(["b (Beta)", f"{slope:.3f}"])  # slope как b
        else:
            report_data.append(["a (Beta)", "Отсутствует"])
            report_data.append(["b (Beta)", "Отсутствует"])

        # Добавляем значения Pn для Gamma
        Ep = [80, 146, 400, 850, 1500, 2515]  # Энергии из gamma_math.py
        if hasattr(self, 'gamma_pn_values') and self.gamma_pn_values:
            for energy, pn_value in zip(Ep, self.gamma_pn_values):
                report_data.append((f"Pn ({energy} кэВ)", f"{pn_value}"))
        else:
            for energy in Ep:
                report_data.append((f"Pn ({energy} кэВ)", "Отсутствует"))

        # Создаём DataFrame
        df = pd.DataFrame(report_data[1:], columns=report_data[0])

        # Создаём и оформляем Excel-файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчёт"

        # Стили
        header_font = Font(bold=True, size=12, color="000000")  # Чёрный шрифт для заголовков
        header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6",
                                  fill_type="solid")  # Голубой фон для заголовков (Light Blue)
        value_font = Font(size=11, color="000000")  # Чёрный шрифт для значений
        beige_fill = PatternFill(start_color="F5F5DC", end_color="F5F5DC",
                                 fill_type="solid")  # Бежевый фон для строк (Beige)
        border = Border(left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin"))

        # Записываем данные
        for r_idx, row in enumerate(df.values, 2):  # Начинаем с 2-й строки, чтобы оставить место для заголовка
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.font = value_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

                # Применяем бежевый фон для всех строк
                cell.fill = beige_fill

                # Если это строка с параметром, содержащим энергию (например, "НУД α, № канала (536.045)"), применяем зелёный фон
                if c_idx == 1 and any(energy in str(value) for energy in
                                      ["536.045", "4385.6", "5687.5", "6192.35", "6337.7", "8044.6"]):
                    for col in range(1, 3):  # Применяем зелёный фон к обеим ячейкам строки
                        ws.cell(row=r_idx, column=col).fill = beige_fill

        # Заголовок
        for c_idx, header in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=c_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # Устанавливаем ширину колонок вручную, как в образце
        ws.column_dimensions['A'].width = 30  # Колонка "Параметр"
        ws.column_dimensions['B'].width = 15  # Колонка "Значение"

        # Сохраняем файл с обработкой ошибки
        try:
            wb.save(report_path)
            # Формируем сообщение с зелёной галочкой
            success_message = f"Отчёт сохранён в файл: {report_path}"
            self.show_info_message(success_message)
            # Выводим в консоль с зелёным цветом
            print(f"\033[92m✔ Отчёт сохранён в файл: {report_path}\033[0m")
        except PermissionError:
            self.show_warning_message(
                "Невозможно создать новый отчёт, пока открыт Excel-файл со старым отчётом. Закройте файл и попробуйте снова.")
            print(f"Ошибка: Невозможно сохранить отчёт в файл {report_path}. Возможно, файл уже открыт в Excel.")
            return
        except Exception as e:
            self.show_warning_message(f"Произошла ошибка при сохранении отчёта: {str(e)}")
            print(f"Ошибка при сохранении отчёта: {str(e)}")
            return

    def open_side_window(self):
        """Открывает дополнительное окно с параметрами."""
        if not hasattr(self, 'side_window') or not self.side_window.isVisible():
            self.side_window = SideWindow(self.geometry(), self.modbus_client, self)
            self.side_window.show()
        else:
            self.side_window.raise_()
            self.side_window.activateWindow()

    def reset_all_data(self):
        """Сбрасывает все данные до начального состояния, как при запуске приложения."""
        # Очистка данных Alfa chart
        self.chart.removeAllSeries()  # Удаляем все серии из графика Alfa
        self.series = QLineSeries()  # Создаем новую пустую серию
        self.series.setName("Импульсы")
        self.series.setPen(QColor("#A3BFFA"))
        self.chart.addSeries(self.series)
        self.series.attachAxis(self.axis_x)
        self.series.attachAxis(self.axis_y)
        self.axis_y.setRange(0, 1)  # Сбрасываем диапазон оси Y

        # Очистка словарей и массивов Alfa
        self.alfa_series_dict.clear()
        self.alfa_checkboxes.clear()
        self.used_alfa_colors.clear()
        self.alfa_data_arrays.clear()
        self.first_impulse_values.clear()
        self.p90_series.clear()
        self.peak_points.clear()
        if hasattr(self, 'original_alfa_series'):
            self.original_alfa_series.clear()

        # Очистка чекбоксов Alfa
        while self.alfa_checkboxes_layout.count():
            item = self.alfa_checkboxes_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        # Очистка данных Beta chart
        self.beta_chart.removeAllSeries()
        self.beta_series = QLineSeries()
        self.beta_series.setName("Beta данные")
        self.beta_chart.addSeries(self.beta_series)
        self.beta_series.attachAxis(self.beta_axis_x)
        self.beta_series.attachAxis(self.beta_axis_y)
        self.beta_axis_y.setRange(0, 1)

        # Очистка словарей и массивов Beta
        self.beta_series_dict.clear()
        self.beta_checkboxes.clear()
        self.used_beta_colors.clear()
        if hasattr(self, 'original_beta_series'):
            self.original_beta_series.clear()
        if hasattr(self, 'cs137_peak_points'):
            self.cs137_peak_points.clear()
        if hasattr(self, 'cs137_peak_coords'):
            self.cs137_peak_coords.clear()

        # Очистка чекбоксов Beta
        while self.checkboxes_layout.count():
            item = self.checkboxes_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        # Очистка данных Gamma chart
        self.gamma_chart.removeAllSeries()
        self.gamma_series = QLineSeries()
        self.gamma_series.setName("Gamma данные")
        self.gamma_chart.addSeries(self.gamma_series)
        self.gamma_series.attachAxis(self.gamma_axis_x)
        self.gamma_series.attachAxis(self.gamma_axis_y)
        self.gamma_axis_y.setRange(0, 1)

        # Очистка словарей и массивов Gamma
        self.gamma_series_dict.clear()
        self.gamma_checkboxes.clear()
        self.used_gamma_colors.clear()
        if hasattr(self, 'original_gamma_series'):
            self.original_gamma_series.clear()

        # Очистка чекбоксов Gamma
        while self.gamma_checkboxes_layout.count():
            item = self.gamma_checkboxes_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        # Очистка массивов данных
        self.fon_data = []
        self.am241_data = []
        self.c14_data = []
        self.cs137_data = []
        self.sry90_data = []
        self.rad_data = []

        # Сброс флагов и калибровочных данных
        self.fon_processed = False
        self.calibration_performed = False
        if hasattr(self, 'calibration_coefficients'):
            delattr(self, 'calibration_coefficients')
        if hasattr(self, 'beta_calibration_coefficients'):
            delattr(self, 'beta_calibration_coefficients')
        if hasattr(self, 'p_values'):
            delattr(self, 'p_values')
        if hasattr(self, 'ra_value'):
            delattr(self, 'ra_value')
        if hasattr(self, 'k1p9_value'):
            delattr(self, 'k1p9_value')
        if hasattr(self, 'k1c0'):
            delattr(self, 'k1c0')

        # Обновление таблицы калибровки
        self.update_calibration_table()

        # Сброс состояния логарифмических чекбоксов
        self.log_checkbox.setChecked(False)
        self.beta_log_checkbox.setChecked(False)
        self.gamma_log_checkbox.setChecked(False)

        # Обновление интерфейса
        self.chart_view.update()
        self.beta_chart_view.update()
        self.gamma_chart_view.update()

    def on_tab_changed(self, index):
        logging.debug(f"Переключение на вкладку с индексом: {index}")
        try:
            # Обновление списка отчётов на вкладке "Combined Report"
            if index == 5:
                self.update_reports_list(index)
            # Обновление таблицы на вкладке "Report"
            elif index == 4:
                self.update_calibration_table()
        except Exception as e:
            logging.error(f"Ошибка в on_tab_changed: {str(e)}")

    ##########################################################################
    # Методы для работы с файлами и контекстным меню
    ##########################################################################
    def auto_load_files(self):
        """
        Эмулирует нажатие на "Загрузить Beta", "Загрузить Alfa" или "Загрузить Gamma" для файлов с определёнными именами.
        Файлы с "gamma" в имени окрашиваются в пурпурный цвет и загружаются на график Gamma,
        кроме файлов "98_fon_2_gamma" и "98_fon_gamma", которые только окрашиваются.
        """
        # Заданная последовательность ключевых слов для Beta (в нижнем регистре)
        beta_keywords = ["фона", "sry90", "rad", "cs137", "c14", "am241"]
        # Заданная последовательность ключевых слов для Alfa (в нижнем регистре)
        alfa_keywords = ["rn", "am241"]
        # Файлы для Gamma (в нижнем регистре)
        gamma_keywords = ["gamma"]

        # Создаём списки файлов для Beta, Alfa и Gamma
        beta_files_to_process = []
        alfa_files_to_process = []
        gamma_files_to_process = []  # Для файлов с "gamma" в имени (кроме исключений)
        gamma_special_files = []  # Для файлов "98_fon_2_gamma" и "98_fon_gamma"

        for i in range(self.file_list.count()):
            item = self.file_list.item(i)
            file_name = item.text().lower()  # Приводим к нижнему регистру для проверки

            # Проверяем специальные гамма-файлы
            if "98_fon_2_gamma" in file_name or "98_fon_gamma" in file_name:
                gamma_special_files.append((file_name, item, i))
                continue  # Пропускаем дальнейшую обработку для этих файлов

            # Проверяем файлы с "gamma" в имени
            for keyword in gamma_keywords:
                if keyword in file_name:
                    gamma_files_to_process.append((keyword, item, i))
                    break

            # Проверяем Beta-файлы (независимо от Gamma)
            for keyword in beta_keywords:
                if keyword in file_name:
                    beta_files_to_process.append((keyword, item, i))
                    break

            # Проверяем Alfa-файлы (независимо от Gamma и Beta)
            for keyword in alfa_keywords:
                if keyword in file_name:
                    alfa_files_to_process.append((keyword, item, i))
                    break

        # Сортируем файлы по порядку ключевых слов
        beta_files_to_process.sort(key=lambda x: beta_keywords.index(x[0]))
        alfa_files_to_process.sort(key=lambda x: alfa_keywords.index(x[0]))
        gamma_files_to_process.sort(key=lambda x: gamma_keywords.index(x[0]))

        # Обрабатываем специальные гамма-файлы (пурпурный цвет, без графика)
        for _, item, _ in gamma_special_files:
            item.setBackground(QColor(147, 112, 219))  # Пурпурный цвет
            print(f"Отмечен пурпурным (без графика): {item.text()}")

        # Обрабатываем остальные гамма-файлы (пурпурный цвет и загрузка на график)
        for keyword, item, row in gamma_files_to_process:
            print(f"Авто загрузка Gamma для файла: {item.text()} (ключевое слово: {keyword})")
            pos = self.file_list.visualItemRect(item).center()
            item.setBackground(QColor(147, 112, 219))  # Пурпурный цвет
            self.change_color(pos, 'gamma')  # Вызываем change_color для загрузки на график
            self.check_color_and_load_data(pos)

        # Эмулируем нажатие на "Загрузить Beta" для Beta-файлов
        for keyword, item, row in beta_files_to_process:
            print(f"Авто загрузка Beta для файла: {item.text()} (ключевое слово: {keyword})")
            pos = self.file_list.visualItemRect(item).center()
            self.change_color(pos, 'beta')
            self.check_color_and_load_data(pos)

        # Эмулируем нажатие на "Загрузить Alfa" для Alfa-файлов
        for keyword, item, row in alfa_files_to_process:
            print(f"Авто загрузка Alfa для файла: {item.text()} (ключевое слово: {keyword})")
            pos = self.file_list.visualItemRect(item).center()
            self.change_color(pos, 'alfa')
            self.check_color_and_load_data(pos)

        # Проверяем, были ли найдены файлы для обработки
        if not beta_files_to_process and not alfa_files_to_process and not gamma_files_to_process and not gamma_special_files:
            self.show_warning_message(
                "Не найдено файлов с именами, содержащими 'фона', 'SrY90', 'Rad', 'Cs137', 'C14', 'Am241' (Beta), 'Rn', 'Am241' (Alfa) или 'gamma' (Gamma)."
            )
        elif not beta_files_to_process:
            self.show_warning_message(
                "Не найдено файлов для Beta с именами, содержащими 'фона', 'SrY90', 'Rad', 'Cs137', 'C14', 'Am241'."
            )
        elif not alfa_files_to_process:
            self.show_warning_message(
                "Не найдено файлов для Alfa с именами, содержащими 'Rn', 'Am241'."
            )
        elif not gamma_files_to_process and not gamma_special_files:
            self.show_warning_message(
                "Не найдено файлов для Gamma с именами, содержащими 'gamma'."
            )
        print_gamma_impulses(self)
        peaks = calculate_peaks(self)

        # Отображаем пики на графике
        if peaks:
            plot_peaks(self, peaks)

    def load_xls_files(self):
        """Загружает список файлов .xls и .xlsx из указанной папки."""
        folder_name = self.folder_input.text()  # Получаем имя папки из поля ввода
        folder_path = os.path.join(os.getcwd(), folder_name)  # Полный путь к папке

        # Если папка изменилась, сбрасываем данные
        if folder_name != self.folder_input.text():
            self.reset_all_data()


        self.file_list.clear()  # Очищаем список перед загрузкой новых файлов

        if os.path.exists(folder_path) and os.path.isdir(folder_path):
            # Получаем список файлов .xls и .xlsx в папке
            xls_files = [f for f in os.listdir(folder_path) if f.endswith((".xls", ".xlsx"))]
            for file_name in xls_files:
                item = QListWidgetItem(file_name)
                self.file_list.addItem(item)
        else:
            # Если папка не существует, выводим сообщение
            self.show_warning_message(f"Папка '{folder_name}' не найдена.")

    def show_context_menu(self, pos):
        """Отображает контекстное меню при правом клике на файл."""
        context_menu = QMenu(self)  # Создаем меню
        open_action = context_menu.addAction("Открыть")  # Добавляем кнопку "Открыть"
        load_alfa_action = context_menu.addAction("Загрузить Alfa")  # Добавляем кнопку "Загрузить Alfa"
        load_beta_action = context_menu.addAction("Загрузить Beta")  # Добавляем кнопку "Загрузить Beta"
        load_gamma_action = context_menu.addAction("Загрузить Gamma")  # Добавляем кнопку "Загрузить Gamma"
        disable_action = context_menu.addAction("Отключить")  # Добавляем кнопку "Отключить"

        # Связываем действия с методами
        open_action.triggered.connect(lambda: self.open_xls_file(pos))
        load_alfa_action.triggered.connect(lambda: self.change_color(pos, 'alfa'))
        load_beta_action.triggered.connect(lambda: self.change_color(pos, 'beta'))
        load_gamma_action.triggered.connect(lambda: self.change_color(pos, 'gamma'))  # Новая опция
        load_alfa_action.triggered.connect(lambda: self.check_color_and_load_data(pos))  # Загрузка данных для Alfa
        load_beta_action.triggered.connect(lambda: self.check_color_and_load_data(pos))  # Загрузка данных для Beta
        load_gamma_action.triggered.connect(lambda: self.check_color_and_load_data(pos))  # Загрузка данных для Gamma
        disable_action.triggered.connect(lambda: self.change_color(pos, 'disable'))

        context_menu.exec(self.file_list.mapToGlobal(pos))  # Показываем меню в точке правого клика

    def change_color(self, pos, action_type):
        """Меняет цвет фона элемента в зависимости от действия и отображает/удаляет графики."""
        index = self.file_list.indexAt(pos)
        item = self.file_list.itemFromIndex(index)

        if item is None:
            return  # Защита от некорректного индекса

        file_name = item.text().lower()  # Приводим к нижнему регистру для проверки
        folder_name = self.folder_input.text()
        folder_path = os.path.join(os.getcwd(), folder_name)
        file_path = os.path.join(folder_path, file_name)

        impulse_value = self.read_first_impulse_value(file_path)
        if impulse_value is not None:
            self.first_impulse_values[file_name] = impulse_value  # Сохраняем значение

        # Проверка на наличие "98_fon_2_gamma" или "98_fon_gamma" в имени файла
        if "98_fon_2_gamma" in file_name or "98_fon_gamma" in file_name:
            item.setBackground(QColor(147, 112, 219))  # Пурпурный цвет
            return  # Не строим график, просто выходим из метода

        # Проверка на наличие "gamma" в имени файла
        if "gamma" in file_name and action_type == 'gamma':
            item.setBackground(QColor(147, 112, 219))  # Пурпурный цвет
            self.add_or_remove_chart(item.text(), 'gamma', True)  # Загружаем на график Gamma
            return

        # Оригинальная логика для Alfa и Beta
        if action_type == 'alfa':
            item.setBackground(QColor(144, 238, 144))  # Салатовый для "Загрузить Alfa"
            self.add_or_remove_chart(item.text(), 'alfa', True)
        elif action_type == 'beta':
            item.setBackground(QColor(173, 216, 230))  # Голубой для "Загрузить Beta"
            self.add_or_remove_chart(item.text(), 'beta', True)
        elif action_type == 'gamma':
            item.setBackground(QColor(147, 112, 219))  # Пурпурный для "Загрузить Gamma"
            self.add_or_remove_chart(item.text(), 'gamma', True)
        elif action_type == 'disable':
            item.setBackground(Qt.GlobalColor.white)  # Обесцвечиваем поле для "Отключить"
            file_name = item.text()
            # Удаляем график и все связанные данные для всех типов
            self.add_or_remove_chart(file_name, 'alfa', False)
            self.add_or_remove_chart(file_name, 'beta', False)
            self.add_or_remove_chart(file_name, 'gamma', False)
            # Очищаем пики и другие зависимости, связанные с Rn
            self.cleanup_rn_data(file_name)

    def read_first_impulse_value(self, file_path):
        """Читает значение 'Кол-во импульсов' из первой строки файла Excel."""
        if not os.path.exists(file_path):
            self.show_warning_message(f"Файл '{os.path.basename(file_path)}' не найден.")
            return None

        try:
            df = pd.read_excel(file_path)
            if 'Кол-во импульсов' not in df.columns:
                self.show_warning_message(
                    f"Столбец 'Кол-во импульсов' не найден в файле '{os.path.basename(file_path)}'.")
                return None

            # Берем значение из первой строки столбца 'Кол-во импульсов'
            first_impulse = df.iloc[0]['Кол-во импульсов']
            return first_impulse

        except Exception as e:
            self.show_error_message(f"Ошибка при чтении данных из файла: {str(e)}")
            return None

    def cleanup_rn_data(self, file_name):
        """Очищает все данные, связанные с графиком Rn."""
        # Удаляем пики из peak_points, если они существуют
        if file_name in self.peak_points:
            peak_series = self.peak_points.pop(file_name, None)
            if peak_series and isinstance(peak_series, dict):
                for peak in peak_series.values():
                    if peak and self.chart.series().contains(peak):
                        self.chart.removeSeries(peak)
            elif peak_series and self.chart.series().contains(peak_series):
                self.chart.removeSeries(peak_series)

        # Пересчитываем линии P90, если график Rn был удален
        if "Rn" in file_name:
            self.highlight_p90_points()  # Обновляем линии P90, если они зависят от Rn

    def save_alfa_data(self, file_name):
        """Сохраняет вторую строку из файла .xls в alfa_data_arrays."""
        folder_name = self.folder_input.text()
        folder_path = os.path.join(os.getcwd(), folder_name)
        file_path = os.path.join(folder_path, file_name)

        if not os.path.exists(file_path):
            self.show_warning_message(f"Файл '{file_name}' не найден.")
            return

        try:
            df = pd.read_excel(file_path)
            if 'Канал' not in df.columns or 'Кол-во импульсов' not in df.columns:
                self.show_warning_message(f"Неверный формат данных в файле '{file_name}'.")
                return
            # Берем только вторую строку (индекс 1)
            if len(df) < 2:
                self.show_warning_message(f"В файле '{file_name}' менее 2 строк.")
                return
            second_row = df.iloc[0]  # Вторая строка
            channel = int(second_row['Канал'])
            impulses = second_row['Кол-во импульсов']
            self.alfa_data_arrays[file_name] = [channel, impulses]  # Сохраняем как список [Канал, Кол-во импульсов]

        except Exception as e:
            self.show_error_message(f"Ошибка при сохранении данных из файла: {str(e)}")

    def add_or_remove_chart(self, file_name, chart_type, add):
        """Добавляет или удаляет график в зависимости от состояния."""
        if chart_type == 'alfa' and not hasattr(self, 'alfa_series_dict'):
            self.alfa_series_dict = {}  # Инициализируем словарь, если его нет
        elif chart_type == 'beta' and not hasattr(self, 'beta_series_dict'):
            self.beta_series_dict = {}  # Инициализируем словарь, если его нет
        elif chart_type == 'gamma' and not hasattr(self, 'gamma_series_dict'):
            self.gamma_series_dict = {}  # Инициализируем словарь, если его нет

        if add:
            if chart_type == 'alfa' and file_name not in self.alfa_series_dict:
                self.load_data_for_chart('alfa', file_name)
            elif chart_type == 'beta' and file_name not in self.beta_series_dict:
                self.load_data_for_chart('beta', file_name)
            elif chart_type == 'gamma' and file_name not in self.gamma_series_dict:
                self.load_data_for_chart('gamma', file_name)
        else:
            if chart_type == 'alfa':
                self.remove_specific_chart('alfa', file_name)
            elif chart_type == 'beta':
                self.remove_specific_chart('beta', file_name)
            elif chart_type == 'gamma':
                self.remove_specific_chart('gamma', file_name)

    def open_xls_file(self, pos):
        """Открывает выбранный файл .xls при правом клике."""
        index = self.file_list.indexAt(pos)
        item = self.file_list.itemFromIndex(index)
        folder_name = self.folder_input.text()  # Получаем имя папки из поля ввода
        folder_path = os.path.join(os.getcwd(), folder_name)  # Полный путь к папке
        file_name = item.text()  # Получаем имя файла из выбранного элемента
        file_path = os.path.join(folder_path, file_name)  # Полный путь к файлу

        if os.path.exists(file_path):
            # Открываем файл с помощью стандартного приложения
            QDesktopServices.openUrl(QUrl.fromLocalFile(file_path))
        else:
            self.show_warning_message(f"Файл '{file_name}' не найден.")

    def remove_specific_chart(self, chart_type, file_name):
        try:
            if chart_type == 'beta':
                if file_name in self.beta_series_dict:
                    series_to_remove = self.beta_series_dict[file_name]
                    self.beta_chart.removeSeries(series_to_remove)

                    # Удаляем точки пиков Cs137, если они существуют
                    if hasattr(self, 'cs137_peak_points'):
                        for peak_point in self.cs137_peak_points:
                            self.beta_chart.removeSeries(peak_point)
                        self.cs137_peak_points.clear()

                    if file_name in self.used_beta_colors:
                        del self.used_beta_colors[file_name]

                    del self.beta_series_dict[file_name]
                    update_calibration_button_state(self)
            elif chart_type == 'gamma':
                if file_name in self.gamma_series_dict:
                    series_to_remove = self.gamma_series_dict[file_name]
                    self.gamma_chart.removeSeries(series_to_remove)

                    if file_name in self.used_gamma_colors:
                        del self.used_gamma_colors[file_name]

                    del self.gamma_series_dict[file_name]
        except Exception as e:
            self.show_error_message(f"Ошибка при удалении графика: {str(e)}")

    ##########################################################################
    # Методы для работы с графиками и данными
    ##########################################################################

    def check_color_and_load_data(self, pos):
        """Проверяет цвет фона элемента и загружает данные из Excel в соответствующий график."""
        index = self.file_list.indexAt(pos)
        item = self.file_list.itemFromIndex(index)

        # Получаем цвет фона элемента
        background_color = item.background().color()

        # Проверяем цвет фона
        if background_color == QColor(144, 238, 144):  # Салатовый цвет для 'Alfa'
            self.load_data_for_chart('alfa', item.text())
        elif background_color == QColor(173, 216, 230):  # Голубой цвет для 'Beta'
            self.load_data_for_chart('beta', item.text())
        elif background_color == QColor(147, 112, 219):  # Пурпурный цвет для 'Gamma'
            self.load_data_for_chart('gamma', item.text())

    def load_data_for_chart(self, chart_type, file_name):
        """Загружает данные из файла Excel и отображает на соответствующем графике."""
        folder_name = self.folder_input.text()  # Получаем имя папки из поля ввода
        folder_path = os.path.join(os.getcwd(), folder_name)  # Полный путь к папке
        file_path = os.path.join(folder_path, file_name)  # Полный путь к файлу

        if not os.path.exists(file_path):
            self.show_warning_message(f"Файл '{file_name}' не найден.")
            return

        try:
            # Чтение данных из Excel
            df = pd.read_excel(file_path)

            # Проверяем структуру данных
            if 'Канал' not in df.columns or 'Кол-во импульсов' not in df.columns:
                self.show_warning_message(f"Неверный формат данных в файле '{file_name}'.")
                return

            # Сохраняем оригинальные значения столбца 'Кол-во импульсов'
            original_impulses = df['Кол-во импульсов'].tolist()

            # Распределяем оригинальные данные по массивам в зависимости от имени файла
            file_name_lower = file_name.lower()  # Приводим имя файла к нижнему регистру для проверки
            if "rad" in file_name_lower:
                self.rad_data = original_impulses
            elif "am241" in file_name_lower:
                self.am241_data = original_impulses
                self.save_alfa_data(file_name)  # Сохраняем данные второй строки для Am241
            elif "c14" in file_name_lower:
                self.c14_data = original_impulses
            elif "sry90" in file_name_lower:
                self.sry90_data = original_impulses
            elif "cs137" in file_name_lower:
                self.cs137_data = original_impulses
            elif "фона" in file_name_lower:
                self.fon_data = original_impulses

            # Заменяем первые три значения импульсов на 0 в копии данных для графика
            df.loc[0:2, 'Кол-во импульсов'] = 0  # Заменяем первые 3 строки на 0 в колонке 'Кол-во импульсов'

            # Отображаем данные на графиках
            if chart_type == 'alfa':
                self.update_alfa_chart(df, file_name, original_impulses)  # Передаём original_impulses
            elif chart_type == 'beta':
                self.update_beta_chart(df, file_name, original_impulses)  # Передаём original_impulses
            elif chart_type == 'gamma':
                self.update_gamma_chart(df, file_name, original_impulses)  # Передаём original_impulses

        except Exception as e:
            self.show_error_message(f"Ошибка при загрузке данных из файла: {str(e)}")

    def update_y_axis_range(self):
        """Пересчитывает диапазон оси Y на основе всех имеющихся серий Alfa."""
        max_y = max((max((point.y() for point in series.points()), default=0)
                     for series in self.alfa_series_dict.values()), default=0)
        self.axis_y.setRange(0, max_y * 1.1)

    def update_beta_y_axis_range(self):
        """Пересчитывает диапазон оси Y на основе всех имеющихся серий Beta."""
        max_y = max((max((point.y() for point in series.points()), default=0)
                     for series in self.beta_series_dict.values()), default=0)
        self.beta_axis_y.setRange(0, max_y * 1.1)

    def adjust_y_axis_for_series(self, series, state):
        """Настраивает ось Y для графика Alfa в зависимости от состояния CheckBox."""
        if series is None or series.count() == 0:
            return

        if state == Qt.CheckState.Checked.value:
            max_y = max(point.y() for point in series.points()) if series.count() > 0 else 0
            self.axis_y.setRange(0, max_y * 1.1)
        else:
            self.update_y_axis_range()

    def adjust_beta_y_axis_for_series(self, series, state):
        """Настраивает ось Y для графика Beta в зависимости от состояния CheckBox."""
        if series is None or series.count() == 0:
            return

        if state == Qt.CheckState.Checked.value:
            max_y = max(point.y() for point in series.points()) if series.count() > 0 else 0
            self.beta_axis_y.setRange(0, max_y * 1.1)
        else:
            self.update_beta_y_axis_range()

    def highlight_p90_points(self):
        """Отмечает вертикальными красными линиями позиции P90 на графике Alfa chart."""
        # Удаляем старые серии, если они существуют
        for series in list(self.p90_series.values()):
            self.chart.removeSeries(series)
        self.p90_series.clear()

        # Используем значения из self.p_values, если они есть
        if not hasattr(self, 'p_values') or not self.p_values:
            print("Предупреждение: Значения P(e) не найдены. Используются значения по умолчанию.")

        else:
            p90_x_values = list(self.p_values.values())  # Используем рассчитанные значения

        # Находим максимальное значение y для установки высоты линий
        max_y = max((max((point.y() for point in series.points()), default=1)
                     for series in self.alfa_series_dict.values()), default=1000)

        # Создаем вертикальные линии для каждой координаты x
        for i, x in enumerate(p90_x_values):
            line_series = QLineSeries()
            energy = list(self.p_values.keys())[i] if hasattr(self, 'p_values') else f"unknown_{x}"
            line_series.setName(f"P({energy})")
            pen = line_series.pen()
            pen.setColor(QColor(255, 0, 0))  # Красный цвет
            pen.setStyle(Qt.PenStyle.DashLine)  # Пунктирный стиль
            line_series.setPen(pen)
            line_series.append(x, 0)  # Начало линии
            line_series.append(x, max_y * 1.1)  # Конец линии (с запасом выше максимума)
            self.chart.addSeries(line_series)
            line_series.attachAxis(self.axis_x)
            line_series.attachAxis(self.axis_y)
            self.p90_series[x] = line_series

    def update_alfa_chart(self, df, file_name, original_data):
        """
        Обновляет график на вкладке Alfa с учетом всех графиков.
        """
        if file_name in self.alfa_series_dict:
            return

        alfa_series = QLineSeries()
        second_word = file_name.split()[1] if len(file_name.split()) > 1 else file_name
        alfa_series.setName(f"({second_word})")

        df_filtered = df[(df['Канал'] >= 200) & (df['Канал'] <= 1023)]
        for _, row in df_filtered.iterrows():
            alfa_series.append(row['Канал'], row['Кол-во импульсов'])

        color = self.get_unique_color(file_name, 'alfa')
        alfa_series.setColor(color)

        self.alfa_series_dict[file_name] = alfa_series
        if not hasattr(self, "original_alfa_series"):
            self.original_alfa_series = {}
        self.original_alfa_series[file_name] = (alfa_series, color)

        self.chart.addSeries(alfa_series)
        alfa_series.attachAxis(self.axis_x)
        alfa_series.attachAxis(self.axis_y)

        highlight_am241_peak(self.chart, alfa_series, self.peak_points)
        highlight_rn_peaks(self.chart, alfa_series, self.peak_points, self)

        if file_name in self.alfa_checkboxes:
            old_checkbox = self.alfa_checkboxes.pop(file_name)
            old_checkbox.setParent(None)
            old_checkbox.deleteLater()

        checkbox = QCheckBox(f"Активировать масштаб для {second_word}")
        checkbox.stateChanged.connect(
            lambda state, series=alfa_series: self.adjust_y_axis_for_series(series, state))
        self.alfa_checkboxes[file_name] = checkbox

        # Добавляем чекбокс в self.alfa_checkboxes_layout
        self.alfa_checkboxes_layout.addWidget(checkbox)

        if not self.alfa_checkboxes_widget.isVisible():
            self.alfa_checkboxes_widget.show()

        self.update_y_axis_range()
        self.highlight_p90_points()

    def update_beta_chart(self, df, file_name, original_data):
        """
        Обновляет график на вкладке Beta с учетом всех графиков.
        """
        if file_name in self.beta_series_dict:
            return

        # Явное преобразование second_word в строку и отладочный вывод
        second_word = str(file_name.split()[1] if len(file_name.split()) > 1 else file_name)
        series_name = f"({second_word})"


        # Проверяем условие с явным приведением к строке
        if "фона" in series_name.lower():
            beta_series = QLineSeries()
            beta_series.setName(series_name)
            for _, row in df.iterrows():
                beta_series.append(row['Канал'], row['Кол-во импульсов'])
            self.beta_series_dict[file_name] = beta_series

            process_fon_data(self, original_data, file_name)
            return

        beta_series = QLineSeries()
        beta_series.setName(series_name)

        # Используем обнулённые данные для отображения
        for _, row in df.iterrows():
            beta_series.append(row['Канал'], row['Кол-во импульсов'])

        # Передаём оригинальные данные для изотопов
        process_isotope_data(self, original_data, file_name)

        color = self.get_unique_color(file_name, 'beta')
        beta_series.setColor(color)

        self.beta_series_dict[file_name] = beta_series
        if not hasattr(self, "original_beta_series"):
            self.original_beta_series = {}
        self.original_beta_series[file_name] = (beta_series, color)

        self.beta_chart.addSeries(beta_series)
        beta_series.attachAxis(self.beta_axis_x)
        beta_series.attachAxis(self.beta_axis_y)

        # Добавляем обработку изотопов
        process_isotope_data(self, original_data, file_name)

        # Выделяем второй пик Cs137
        if "Cs137" in file_name:
            self.highlight_cs137_second_peak()

        if file_name in self.beta_checkboxes:
            old_checkbox = self.beta_checkboxes.pop(file_name)
            old_checkbox.setParent(None)
            old_checkbox.deleteLater()

        checkbox = QCheckBox(f"Активировать масштаб для {second_word}")
        checkbox.stateChanged.connect(
            lambda state, series=beta_series: self.adjust_beta_y_axis_for_series(series, state))
        self.beta_checkboxes[file_name] = checkbox

        # Добавляем чекбокс в self.checkboxes_layout
        self.checkboxes_layout.addWidget(checkbox)

        self.update_beta_y_axis_range()
        update_calibration_button_state(self)

    def update_gamma_chart(self, df, file_name, original_data):
        """
        Обновляет график на вкладке Gamma с учетом всех графиков, игнорируя первое значение Y.
        """
        if file_name in self.gamma_series_dict:
            return

        gamma_series = QLineSeries()
        second_word = file_name.split()[1] if len(file_name.split()) > 1 else file_name
        gamma_series.setName(f"({second_word})")

        # Используем обнулённые данные для отображения, начиная со второго значения (индекс 1)
        for index, row in df.iterrows():
            if index == 0:  # Пропускаем первое значение
                continue
            gamma_series.append(row['Канал'], row['Кол-во импульсов'])

        color = self.get_unique_color(file_name, 'gamma')
        gamma_series.setColor(color)

        self.gamma_series_dict[file_name] = gamma_series
        if not hasattr(self, "original_gamma_series"):
            self.original_gamma_series = {}
        self.original_gamma_series[file_name] = (gamma_series, color)

        self.gamma_chart.addSeries(gamma_series)
        gamma_series.attachAxis(self.gamma_axis_x)
        gamma_series.attachAxis(self.gamma_axis_y)

        if file_name in self.gamma_checkboxes:
            old_checkbox = self.gamma_checkboxes.pop(file_name)
            old_checkbox.setParent(None)
            old_checkbox.deleteLater()

        checkbox = QCheckBox(f"Активировать масштаб для {second_word}")
        checkbox.stateChanged.connect(
            lambda state, series=gamma_series: self.adjust_gamma_y_axis_for_series(series, state))
        self.gamma_checkboxes[file_name] = checkbox

        # Добавляем чекбокс в self.gamma_checkboxes_layout
        self.gamma_checkboxes_layout.addWidget(checkbox)

        if not self.gamma_checkboxes_widget.isVisible():
            self.gamma_checkboxes_widget.show()

        self.update_gamma_y_axis_range()

    def get_unique_color(self, file_name, chart_type):
        """Выбирает уникальный цвет для новой серии, избегая повторений."""
        if chart_type == 'alfa':
            used_colors = self.used_alfa_colors
            available_colors = self.available_colors.copy()
        elif chart_type == 'beta':
            used_colors = self.used_beta_colors
            available_colors = self.available_colors.copy()
        elif chart_type == 'gamma':
            used_colors = self.used_gamma_colors
            available_colors = self.available_colors.copy()
        else:
            return QColor(0, 0, 0)  # Чёрный по умолчанию

        # Удаляем уже использованные цвета из доступных
        for color in used_colors.values():
            if color in available_colors:
                available_colors.remove(color)

        if not available_colors:  # Если все цвета использованы, сбрасываем или генерируем новые
            available_colors = self.available_colors.copy()  # Сбрасываем к начальному набору

        # Выбираем первый доступный цвет
        color = available_colors[0]
        used_colors[file_name] = color  # Сохраняем цвет для этого файла
        return color

    def highlight_cs137_second_peak(self):
        """
        Обнаруживает второй пик Cs137 динамически, находит его наивысшую точку и выделяет её точкой на графике.
        Сохраняет координаты точки для корректного отображения в логарифмическом и линейном масштабе.
        """
        if not self.cs137_data or len(self.cs137_data) < 10:
            self.show_warning_message("Недостаточно данных Cs137 для анализа пиков.")
            return

        cs137_data = np.array(self.cs137_data)

        # 1. Сглаживание с помощью скользящего среднего
        window_size = min(5, len(cs137_data) - 1)
        if window_size % 2 == 0:
            window_size += 1
        smoothed_data = np.convolve(cs137_data, np.ones(window_size) / window_size, mode='same')

        # 2. Обнаружение первого пика (глобальный максимум)
        max_peak_idx = np.argmax(smoothed_data)
        max_peak_value = smoothed_data[max_peak_idx]
        threshold = max_peak_value * 0.1
        start_second_peak = max_peak_idx
        for i in range(max_peak_idx, len(smoothed_data)):
            if smoothed_data[i] < threshold:
                start_second_peak = i
                break
        if start_second_peak >= len(smoothed_data) - 1:
            self.show_warning_message("Не удалось определить начало второго пика.")
            return

        # 3. Поиск второго пика в оставшемся диапазоне
        remaining_data = smoothed_data[start_second_peak:]
        peaks = []
        for i in range(1, len(remaining_data) - 1):
            if remaining_data[i] > remaining_data[i - 1] and remaining_data[i] > remaining_data[i + 1]:
                peaks.append(i + start_second_peak)

        if not peaks:
            self.show_warning_message("Второй пик Cs137 не обнаружен.")
            return

        second_peak_idx = peaks[0]
        second_peak_x = second_peak_idx
        second_peak_y = cs137_data[second_peak_x]

        # 4. Уточняем наивысшую точку в диапазоне ±10
        peak_range_start = max(0, second_peak_x - 10)
        peak_range_end = min(len(cs137_data) - 1, second_peak_x + 10)
        peak_range_data = cs137_data[peak_range_start:peak_range_end + 1]
        refined_peak_x = peak_range_start + np.argmax(peak_range_data)
        refined_peak_y = cs137_data[refined_peak_x]

        print(f"Второй пик Cs137 обнаружен: x={refined_peak_x}, y={refined_peak_y}")

        # Сохраняем координаты пика
        if not hasattr(self, 'cs137_peak_coords'):
            self.cs137_peak_coords = []
        self.cs137_peak_coords.append((refined_peak_x, refined_peak_y))

        # 5. Выделяем точку на графике
        self.draw_cs137_peak_point()

    def draw_cs137_peak_point(self):
        """
        Рисует точку пика Cs137 на графике с учётом текущего масштаба.
        """
        # Удаляем предыдущие точки, если они есть
        if hasattr(self, 'cs137_peak_points'):
            for peak_point in self.cs137_peak_points:
                self.beta_chart.removeSeries(peak_point)
            self.cs137_peak_points.clear()

        if not hasattr(self, 'cs137_peak_coords') or not self.cs137_peak_coords:
            return

        # Проверяем текущий масштаб
        is_log_scale = isinstance(self.beta_axis_y, QLogValueAxis)

        # Создаём новую точку
        peak_point = QScatterSeries()
        peak_point.setName(f"Второй пик Cs137 (x={self.cs137_peak_coords[0][0]})")
        peak_point.setColor(QColor(255, 0, 0))
        peak_point.setMarkerSize(10.0)

        # Получаем координаты
        peak_x, peak_y = self.cs137_peak_coords[0]

        # Если логарифмический масштаб, корректируем y (y >= 1 для QLogValueAxis)
        if is_log_scale:
            peak_y = max(peak_y, 1.0)  # Логарифмическая ось не допускает значений < 1
        peak_point.append(peak_x, peak_y)

        self.beta_chart.addSeries(peak_point)
        peak_point.attachAxis(self.beta_axis_x)
        peak_point.attachAxis(self.beta_axis_y)

        if not hasattr(self, 'cs137_peak_points'):
            self.cs137_peak_points = []
        self.cs137_peak_points.append(peak_point)

    ##########################################################################
    # Методы для работы с графиками и масштабированием
    ##########################################################################

    def toggle_gamma_log_scale(self, state):
        if state == Qt.CheckState.Checked.value:
            # Логарифмический масштаб
            self.gamma_axis_y = QLogValueAxis()
            self.gamma_axis_y.setTitleText("Импульсы (лог)")
            self.gamma_axis_y.setBase(10.0)
            self.gamma_axis_y.setMinorTickCount(9)
            axis_font = QFont("Montserrat", 12)
            self.gamma_axis_y.setLabelsFont(axis_font)
            self.gamma_axis_y.setTitleFont(axis_font)
            self.gamma_axis_y.setTitleBrush(QColor("#000000"))
            self.gamma_axis_y.setLabelsColor(QColor("#000000"))
            self.gamma_axis_y.setGridLineColor(QColor("#F5F5F5"))
            self.gamma_chart.removeAxis(self.gamma_chart.axisY())
            self.gamma_chart.addAxis(self.gamma_axis_y, Qt.AlignmentFlag.AlignLeft)
            self.gamma_series.attachAxis(self.gamma_axis_y)
        else:
            # Линейный масштаб
            self.gamma_axis_y = QValueAxis()
            self.gamma_axis_y.setTitleText("Импульсы")
            axis_font = QFont("Montserrat", 12)
            self.gamma_axis_y.setLabelsFont(axis_font)
            self.gamma_axis_y.setTitleFont(axis_font)
            self.gamma_axis_y.setTitleBrush(QColor("#000000"))
            self.gamma_axis_y.setLabelsColor(QColor("#000000"))
            self.gamma_axis_y.setGridLineColor(QColor("#F5F5F5"))
            self.gamma_chart.removeAxis(self.gamma_chart.axisY())
            self.gamma_chart.addAxis(self.gamma_axis_y, Qt.AlignmentFlag.AlignLeft)
            self.gamma_series.attachAxis(self.gamma_axis_y)

    def toggle_log_scale(self, state):
        """Переключает между логарифмическим и линейным масштабом для всех серий Alfa."""
        if state == Qt.CheckState.Checked.value:
            self.apply_log_scale()
        else:
            self.apply_linear_scale()
        # Перерисовываем линии P90 после смены масштаба
        self.highlight_p90_points()

    def toggle_beta_log_scale(self, state):
        """Переключает между логарифмическим и линейным масштабом для всех серий Beta."""
        if state == Qt.CheckState.Checked.value:
            self.apply_beta_log_scale()
        else:
            self.apply_beta_linear_scale()

    def apply_log_scale(self):
        """Применяет логарифмический масштаб ко всем сериям Alfa с началом оси Y от 1."""
        if not hasattr(self, "original_alfa_series"):
            self.original_alfa_series = {}

        # Создаем логарифмическую ось Y
        log_axis_y = QLogValueAxis()
        log_axis_y.setTitleText("Значение импульса (логарифмический масштаб)")
        log_axis_y.setBase(10.0)
        log_axis_y.setMinorTickCount(9)  # Увеличиваем количество делений для точности

        # Удаляем старую ось Y
        self.chart.removeAxis(self.axis_y)

        # Удаляем все текущие серии из графика
        for series in self.alfa_series_dict.values():
            self.chart.removeSeries(series)

        # Создаем новые логарифмические серии
        new_series_dict = {}
        global_max_y = float('-inf')

        # Минимальное значение для логарифма (ниже 1 не допускаем)
        min_y_threshold = 1.0

        # Обрабатываем данные для каждой серии
        for file_name, (original_series, color) in self.original_alfa_series.items():
            log_series = QLineSeries()
            log_series.setName(original_series.name())
            log_series.setColor(color)  # Сохраняем цвет оригинальной серии

            # Преобразуем данные для логарифмического масштаба
            for point in original_series.points():
                x, y = point.x(), point.y()
                # Устанавливаем минимальное значение y как 1 для корректного отображения
                y_log = max(y, min_y_threshold)
                log_series.append(x, y_log)
                global_max_y = max(global_max_y, y_log)

            new_series_dict[file_name] = log_series
            self.chart.addSeries(log_series)
            log_series.attachAxis(self.axis_x)

        # Устанавливаем диапазон логарифмической оси, начиная с 1
        log_axis_y.setRange(min_y_threshold, global_max_y * 1.5)  # Фиксируем минимум на 1

        # Добавляем новую ось и привязываем серии
        self.chart.addAxis(log_axis_y, Qt.AlignmentFlag.AlignLeft)
        for series in new_series_dict.values():
            series.attachAxis(log_axis_y)

        # Обновляем словари и текущую ось
        self.alfa_series_dict = new_series_dict
        self.axis_y = log_axis_y

        # Перерисовываем пики
        self.reapply_peaks()

    def apply_linear_scale(self):
        """Применяет линейный масштаб ко всем сериям Alfa."""
        if not hasattr(self, "original_alfa_series"):
            return

        # Создаем линейную ось Y
        linear_axis_y = QValueAxis()
        linear_axis_y.setTitleText("Значение импульса")

        # Удаляем старую ось Y
        self.chart.removeAxis(self.axis_y)

        # Удаляем все текущие серии
        for series in self.alfa_series_dict.values():
            self.chart.removeSeries(series)

        # Восстанавливаем оригинальные серии с их цветами
        self.alfa_series_dict = {}
        global_max_y = float('-inf')

        for file_name, (original_series, color) in self.original_alfa_series.items():
            original_series.setColor(color)  # Убеждаемся, что цвет сохранён
            self.alfa_series_dict[file_name] = original_series
            self.chart.addSeries(original_series)
            original_series.attachAxis(self.axis_x)
            y_values = [point.y() for point in original_series.points()]
            if y_values:
                global_max_y = max(global_max_y, max(y_values))

        # Устанавливаем диапазон линейной оси
        linear_axis_y.setRange(0, global_max_y * 1.1 if global_max_y > 0 else 1)
        self.chart.addAxis(linear_axis_y, Qt.AlignmentFlag.AlignLeft)
        for series in self.alfa_series_dict.values():
            series.attachAxis(linear_axis_y)

        # Обновляем текущую ось
        self.axis_y = linear_axis_y

        # Перерисовываем пики
        self.reapply_peaks()

    def apply_beta_log_scale(self):
        """
        Применяет логарифмический масштаб ко всем сериям Beta с началом оси Y от 1.
        Перерисовывает точку пика Cs137.
        """
        if not hasattr(self, "original_beta_series"):
            self.original_beta_series = {}

        # Создаём логарифмическую ось Y
        beta_log_axis_y = QLogValueAxis()
        beta_log_axis_y.setTitleText("Значение импульса (логарифмический масштаб)")
        beta_log_axis_y.setBase(10.0)
        beta_log_axis_y.setMinorTickCount(9)

        # Удаляем старую ось Y
        self.beta_chart.removeAxis(self.beta_axis_y)

        # Удаляем все текущие серии из графика
        for series in self.beta_series_dict.values():
            self.beta_chart.removeSeries(series)

        # Создаём новые логарифмические серии
        new_series_dict = {}
        global_max_y = float('-inf')
        min_y_threshold = 1.0

        for file_name, (original_series, color) in self.original_beta_series.items():
            log_series = QLineSeries()
            log_series.setName(original_series.name())
            log_series.setColor(color)

            for point in original_series.points():
                x, y = point.x(), point.y()
                y_log = max(y, min_y_threshold)
                log_series.append(x, y_log)
                global_max_y = max(global_max_y, y_log)

            new_series_dict[file_name] = log_series
            self.beta_chart.addSeries(log_series)
            log_series.attachAxis(self.beta_axis_x)

        beta_log_axis_y.setRange(min_y_threshold, global_max_y * 1.5)
        self.beta_chart.addAxis(beta_log_axis_y, Qt.AlignmentFlag.AlignLeft)
        for series in new_series_dict.values():
            series.attachAxis(beta_log_axis_y)

        self.beta_series_dict = new_series_dict
        self.beta_axis_y = beta_log_axis_y

        # Перерисовываем точку пика
        self.draw_cs137_peak_point()

    def apply_beta_linear_scale(self):
        """
        Применяет линейный масштаб ко всем сериям Beta.
        Перерисовывает точку пика Cs137.
        """
        if not hasattr(self, "original_beta_series"):
            return

        # Создаём линейную ось Y
        beta_linear_axis_y = QValueAxis()
        beta_linear_axis_y.setTitleText("Значение")

        # Удаляем старую ось Y
        self.beta_chart.removeAxis(self.beta_axis_y)

        # Удаляем все текущие серии
        for series in self.beta_series_dict.values():
            self.beta_chart.removeSeries(series)

        # Восстанавливаем оригинальные серии
        self.beta_series_dict = {}
        global_max_y = float('-inf')

        for file_name, (original_series, color) in self.original_beta_series.items():
            original_series.setColor(color)
            self.beta_series_dict[file_name] = original_series
            self.beta_chart.addSeries(original_series)
            original_series.attachAxis(self.beta_axis_x)
            y_values = [point.y() for point in original_series.points()]
            if y_values:
                global_max_y = max(global_max_y, max(y_values))

        beta_linear_axis_y.setRange(0, global_max_y * 1.1 if global_max_y > 0 else 1)
        self.beta_chart.addAxis(beta_linear_axis_y, Qt.AlignmentFlag.AlignLeft)
        for series in self.beta_series_dict.values():
            series.attachAxis(beta_linear_axis_y)

        self.beta_axis_y = beta_linear_axis_y

        # Перерисовываем точку пика
        self.draw_cs137_peak_point()

    def reapply_peaks(self):
        """Перерисовывает пики после смены масштаба."""
        # Удаляем старые пики
        for peak_data in self.peak_points.values():
            if isinstance(peak_data, dict):
                for peak_series in peak_data.values():
                    if peak_series in self.chart.series():
                        self.chart.removeSeries(peak_series)
            elif peak_data in self.chart.series():
                self.chart.removeSeries(peak_data)

        # Переприменяем пики для всех серий
        self.peak_points.clear()
        for series in self.alfa_series_dict.values():
            highlight_am241_peak(self.chart, series, self.peak_points)
            highlight_rn_peaks(self.chart, series, self.peak_points, self)

    ##########################################################################
    # Методы для работы Gamma
    ##########################################################################

    def toggle_gamma_log_scale(self, state):
        """Переключает между логарифмическим и линейным масштабом для всех серий Gamma."""
        if state == Qt.CheckState.Checked.value:
            self.apply_gamma_log_scale()
        else:
            self.apply_gamma_linear_scale()

    def reapply_gamma_peaks(self):
        """Перерисовывает пики на графике Gamma после смены масштаба или зума."""
        if hasattr(self, 'gamma_peaks'):
            plot_peaks(self, self.gamma_peaks)

    def apply_gamma_log_scale(self):
        """Применяет логарифмический масштаб ко всем сериям Gamma с началом оси Y от 1."""
        if not hasattr(self, "original_gamma_series"):
            self.original_gamma_series = {}

        gamma_log_axis_y = QLogValueAxis()
        gamma_log_axis_y.setTitleText("Импульсы (логарифмический масштаб)")
        gamma_log_axis_y.setBase(10.0)
        gamma_log_axis_y.setMinorTickCount(9)

        self.gamma_chart.removeAxis(self.gamma_axis_y)
        for series in self.gamma_series_dict.values():
            self.gamma_chart.removeSeries(series)

        new_series_dict = {}
        global_max_y = float('-inf')
        min_y_threshold = 1.0

        for file_name, (original_series, color) in self.original_gamma_series.items():
            log_series = QLineSeries()
            log_series.setName(original_series.name())
            log_series.setColor(color)

            for point in original_series.points():
                x, y = point.x(), point.y()
                y_log = max(y, min_y_threshold)
                log_series.append(x, y_log)
                global_max_y = max(global_max_y, y_log)

            new_series_dict[file_name] = log_series
            self.gamma_chart.addSeries(log_series)
            log_series.attachAxis(self.gamma_axis_x)

        gamma_log_axis_y.setRange(min_y_threshold, global_max_y * 1.5)
        self.gamma_chart.addAxis(gamma_log_axis_y, Qt.AlignmentFlag.AlignLeft)
        for series in new_series_dict.values():
            series.attachAxis(gamma_log_axis_y)

        self.gamma_series_dict = new_series_dict
        self.gamma_axis_y = gamma_log_axis_y

        # Перерисовываем пики
        self.reapply_gamma_peaks()

    def apply_gamma_linear_scale(self):
        """Применяет линейный масштаб ко всем сериям Gamma."""
        if not hasattr(self, "original_gamma_series"):
            return

        gamma_linear_axis_y = QValueAxis()
        gamma_linear_axis_y.setTitleText("Импульсы")

        self.gamma_chart.removeAxis(self.gamma_axis_y)
        for series in self.gamma_series_dict.values():
            self.gamma_chart.removeSeries(series)

        self.gamma_series_dict = {}
        global_max_y = float('-inf')

        for file_name, (original_series, color) in self.original_gamma_series.items():
            original_series.setColor(color)
            self.gamma_series_dict[file_name] = original_series
            self.gamma_chart.addSeries(original_series)
            original_series.attachAxis(self.gamma_axis_x)
            y_values = [point.y() for point in original_series.points()]
            if y_values:
                global_max_y = max(global_max_y, max(y_values))

        gamma_linear_axis_y.setRange(0, global_max_y * 1.1 if global_max_y > 0 else 1)
        self.gamma_chart.addAxis(gamma_linear_axis_y, Qt.AlignmentFlag.AlignLeft)
        for series in self.gamma_series_dict.values():
            series.attachAxis(gamma_linear_axis_y)

        self.gamma_axis_y = gamma_linear_axis_y

        # Перерисовываем пики
        self.reapply_gamma_peaks()

    def reset_gamma_zoom(self):
        """Сбрасывает масштаб графика Gamma до исходного состояния."""
        self.gamma_axis_x.setRange(500, 800)
        if self.gamma_series_dict:
            max_y = max(
                max(series.points(), key=lambda p: p.y()).y()
                for series in self.gamma_series_dict.values()
                if series.points()
            )
            self.gamma_axis_y.setRange(0, max_y * 1.1)
        else:
            self.gamma_axis_y.setRange(0, 1)
        self.gamma_chart_view.update()

    def toggle_gamma_checkboxes(self):
        """Переключает видимость виджета с чекбоксами для Gamma chart и изменяет размер окна."""
        if self.gamma_checkboxes_widget.isVisible():
            # Скрываем виджет с чекбоксами
            self.gamma_checkboxes_widget.hide()
            self.gamma_toggle_checkboxes_button.setText("📋 Чекбоксы")
            # Возвращаем исходный размер окна
            self.resize(self.width(), self.original_height)
        else:
            # Показываем виджет с чекбоксами
            self.gamma_checkboxes_widget.show()
            self.gamma_toggle_checkboxes_button.setText("📋 Скрыть")
            # Увеличиваем высоту окна на размер виджета с чекбоксами
            checkboxes_height = self.gamma_checkboxes_widget.sizeHint().height()
            new_height = self.height() + checkboxes_height
            self.resize(self.width(), new_height)

    def update_gamma_y_axis_range(self):
        """Пересчитывает диапазон оси Y на основе всех имеющихся серий Gamma."""
        max_y = max((max((point.y() for point in series.points()), default=0)
                     for series in self.gamma_series_dict.values()), default=0)
        self.gamma_axis_y.setRange(0, max_y * 1.1)

    def adjust_gamma_y_axis_for_series(self, series, state):
        """Настраивает ось Y для графика Gamma в зависимости от состояния CheckBox."""
        if series is None or series.count() == 0:
            return

        if state == Qt.CheckState.Checked.value:
            max_y = max(point.y() for point in series.points()) if series.count() > 0 else 0
            self.gamma_axis_y.setRange(0, max_y * 1.1)
        else:
            self.update_gamma_y_axis_range()

    def perform_calibration(self):
        perform_calibration(self)

    ##########################################################################
    # Методы для работы с экспортом данных и сообщениями
    ##########################################################################

    def show_info_message(self, message):
        """Показывает информационное сообщение."""
        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Icon.Information)
        msg_box.setWindowTitle("Информация")
        msg_box.setText(message)
        msg_box.exec()

    def show_warning_message(self, message):
        """Показывает предупреждение."""
        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Icon.Warning)
        msg_box.setWindowTitle("Предупреждение")
        msg_box.setText(message)
        msg_box.exec()

    def show_error_message(self, message):
        """Показывает сообщение об ошибке."""
        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Icon.Critical)
        msg_box.setWindowTitle("Ошибка")
        msg_box.setText(message)
        msg_box.exec()

##########################################################################
# Основной блок запуска приложения
##########################################################################

if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setWindowIcon(QIcon("app_icon.png"))

    dialog = SettingsDialog()
    if dialog.exec() == QDialog.DialogCode.Rejected:
        sys.exit(0)

    settings = dialog.get_settings()

    splash = QSplashScreen(QPixmap("lib\\Pictures\\Micasensor.png"))
    splash.showMessage("Загрузка...", Qt.AlignmentFlag.AlignBottom | Qt.AlignmentFlag.AlignHCenter,
                       Qt.GlobalColor.white)
    splash.show()

    app.processEvents()
    modbus = ModbusClient(
        port=settings["port"],
        baud_rate=settings["baud_rate"],
        parity=settings["parity"],
        stop_bits=settings["stop_bits"],
        byte_size=settings["byte_size"],
        timeout=settings["timeout"]
    )

    QTimer.singleShot(2000, splash.close)

    window = SpectrumWindow(modbus)
    window.resize(800, 600)
    window.original_height = 600
    # Получаем доступную геометрию экрана
    screen_geometry = app.primaryScreen().availableGeometry()
    # Перемещаем окно к верхнему левому углу (x=0 для левого края, y=0 для верхнего края)
    x_position = screen_geometry.x() + (screen_geometry.width() - window.width()) // 2
    y_position = screen_geometry.y()
    window.move(x_position, y_position)
    QTimer.singleShot(2000, window.show)

    sys.exit(app.exec())